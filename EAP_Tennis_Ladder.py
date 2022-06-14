### ΕΙΣΑΓΩΓΗ ###
import tkinter  
import tkinter.ttk
from tkinter import *
from tkinter import messagebox
import sys
from sys import *
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from configparser import ConfigParser
import os
import os.path
import datetime
from datetime import date,timedelta,datetime
from PIL import ImageTk, Image
from PIL import *
import pywintypes
import win32api



###########################ΜΕΤΑΒΛΗΤΕΣ ΠΡΟΓΡΑΜΜΑΤΟΣ####################################################
global WILDCARD  # ΔΗΛΩΝΕΙ ΠΟΣΑ ΜΑΤΣ ΠΡΕΠΕΙ ΝΑ ΠΑΙΞΕΙ Ο ΠΑΙΧΤΗΣ ΓΙΑ ΝΑ ΠΑΡΕΙ ΜΠΑΛΑΝΤΕΡ
WILDCARD = 5
global MAX_ACTIVE_CHALLENGES  # ΔΗΛΩΝΕΙ ΤΟΝ ΜΕΓΙΣΤΟ ΑΡΙΘΜΟ ΠΡΟΚΛΗΣΕΩΝ ΠΟΥ ΜΠΟΡΕΙ ΝΑ EXEI ENEPΓΕΣ  ΕΝΑΣ ΠΑΙΧΤΗΣ
MAX_ACTIVE_CHALLENGES = 1
global MAX_RANKING_CHALLENGE  # ΔΗΛΩΝΕΙ ΤΗΝ ΜΕΓΙΣΤΗ ΑΠΟΣΤΑΣΗ(ΔΙΑΦΟΡΑ ΚΑΤΑΤΑΞΗΣ) ΠΟΥ ΠΡΕΠΕΙ ΝΑ ΕΧΟΥΝΕ ΟΙ ΠΑΙΚΤΕΣ ΓΙΑ ΝΑ ΣΤΕΙΛΟΥΝΕ ΠΡΟΚΛΗΣΗ
MAX_RANKING_CHALLENGE = 5
global POINT_SET_MATCH # ΟΡΙΖΕΙ ΤΟΝ ΜΕΓΙΣΤΟ ΑΡΙΘΜΟ ΣΕΤ ΠΟΥ ΠΡΕΠΕΙ ΝΑ ΠΑΡΕΙ ΕΝΑΣ ΠΑΙΚΤΗΣ ΓΙΑ ΝΑ ΚΕΡΔΙΣΕΙ ΤΟ ΜΑΤΣ ΠΡΟΚΛΗΣΗΣ
POINT_SET_MATCH = 3
global filename_ranking # ΔΗΛΩΝΕΙ ΤΟ ΑΡΧΕΙΟ ΕXCEL ΠΟΥ ΣΩΖΟΝΤΑΙ ΤΑ ΣΤΟΙΧΕΙΑ ΤΩΝ ΠΑΙΚΤΩΝ
filename_ranking = os.path.abspath("ranking.xlsx") # ΤΟ ΑΡΧΕΙΟ ΜΕ ΤΗΝ ΚΑΤΑΤΑΞΗ
filename_conf = os.path.abspath("config.ini") # ΤΟ ΑΡΧΕΙΟ ΡΥΘΜΙΣΕΩΝ
global filename_image # δηλωνει το αρχειο εικονας για το login page
filename_image = os.path.abspath('background.png')
#########################################################################################################
pl_data = [] # ΛΙΣΤΑ ΜΕ ΤΑ ΣΤΟΙΧΕΙΑ ΤΩΝ ΠΑΙΚΤΩΝ
ch_out = [] # ΛΙΣΤΑ ΜΕ ΤΙΣ ΜΗ ΕΠΙΒΕΒΑΙΩΜΕΝΕΣ ΠΡΟΚΛΗΣΕΙΣ
ch_valid = [] # ΛΙΣΤΑ ΜΕ ΤΙΣ ΕΓΚΥΡΕΣ ΠΡΟΚΛΗΣΕΙΣ
statistics = [] # ΛΙΣΤΑ ΠΟΥ ΑΠΟΘΗΚΕΥΕΙ ΤΑ ΣΤΑΤΙΣΤΙΚΑ ΤΩΝ ΠΑΙΓΜΕΝΩΝ ΠΑΙΧΝΙΔΙΩΝ
#########################################################################################################
##########################ΟΡΙΣΜΟΣ ΚΛΑΣΗΣ ΠΑΙΚΤΗΣ########################################################
class Player:
    def __init__(self,name,surname,age,m_pl,m_won,m_lost,sets_pl,sets_won,sets_lost,wildcard,active_ch_counter,active_ch):
        global WILDCARD
        global MAX_ACTIVE_CHALLENGES
        self.name = name    #ΟΝΟΜΑ
        self.surname = surname # ΕΠΩΝΥΜΟ
        self.age = age #ΗΛΙΚΙΑ
        self.m_pl = m_pl # ΜΑΤΣ ΣΥΝΟΛΟ
        self.m_won = m_won # ΚΕΡΔΙΣMΕΝΑ ΜΑΤΣ
        self.m_lost = m_lost # ΧΑΜΕΝΑ ΜΑΤΣ
        self.sets_pl = sets_pl # ΣΕΤ ΣΥΝΟΛΟ
        self.sets_won = sets_won # ΚΕΡΔΙΣMΕΝΑ ΣΕΤ
        self.sets_lost = sets_lost # ΧΑΜΕΝΑ ΣΕΤ
        self.wildcard = wildcard # ΤΟ ΑΝ ΕΧΕΙ ΜΠΑΛΑΝΤΕΡ
        self.active_ch_counter = active_ch_counter # ΑΡΙΘΜΟΣ ΠΡΟΚΛΗΣΕΩΝ ΔΕΝ ΜΠΟΡΕΙ ΑΝ ΥΠΕΡΒΕΙ ΤΟ MAX_ΑCTIVE_CHALLENGES
        self.active_ch = active_ch # ΤΟ ΕΑΝ ΜΠΟΡΕΙ ΝΑ ΔΕΚΤΕΙ Η ΝΑ ΔΩΣΕΙ ΠΡΟΚΛΗΣH
        return
    def increase_m_pl(self): # ΑΥΞΑΝΕΙ ΤΑ ΜΑΤΣ ΠΡΟΚΛΗΣΗΣ
        self.m_pl = self.m_pl+1
        return
    def increase_m_won(self): # ΑΥΞΑΝΕΙ ΤΑ ΜΑΤΣ ΠΟΥ ΠΑΙΧΤΗΚΑΝ ΣΥΝΟΛΙΚΑ
        self.m_won = self.m_won +1 
        return
    def increase_m_lost(self):#ΑΥΞΑΝΕΙ ΤΑ ΜΑΤΣ ΠΟΥ ΧΑΘΗΚΑΝ
        self.m_lost = self.m_lost +1
        return
    def increase_sets_pl(self):# ΑΥΞΑΝΕΙ ΤΑ ΣΕΤΣ ΠΟΥ ΠΑΙΧΤΗΚΑΝ ΣΥΝΟΛΙΚΑ
        self.sets_pl = self.sets_won + self.sets_lost
        return
    def increase_sets_won(self,number): # ΑΥΞΑΝΕΙ ΤΑ ΣΕΤΣ ΠΟΥ ΚΕΡΔΙΘΗΚΑΝ
        self.sets_won = self.sets_won + number
        return
    def increase_sets_lost(self,number): # ΑΥΞΑΝΕΙ ΤΑ ΣΕΤΣ ΠΟΥ ΧΑΘΗΚΑΝ
        self.sets_lost = self.sets_lost + number
        return
    def change_wild_state(self): # ΑΛΛΑΖΕΙ ΤΗΝ ΚΑΤΑΣΤΑΣΗ ΤΟΥ WILDCARD TRUE <--> FALSE
        if(self.wildcard==False):
            self.wildcard = True
        else:
            self.wildcard = False
        return
    def check_wild_state(self): # EΛΕΓΧΕΙ ΑΝ ΤΟ WILDCARD ΠΡΕΠΕΙ ΑΝ ΑΛΛΑΧΘΕΙ ΚΑΙ ΚΑΛΕΙ ΤΗΝ ΚΑΤΑΛΛΗΛΗ ΣΥΝΑΡΤΗΣΗ
        if(self.m_pl>=WILDCARD and self.m_pl % WILDCARD==0):
            self.change_wild_state()
        return  
    def increase_active_ch_counter(self): # ΑΥΞΑΝΕΙ ΤΟΝ ΑΡΙΘΜΟ ΤΩΝ ΠΡΟΚΛΗΣΕΩΝ ΠΟΥ ΕΧΕΙ ΔΩΣΕΙ Η ΔΕΚΤΕΙ Ο ΠΑΙΚΤΗΣ
        self.active_ch_counter = self.active_ch_counter + 1
        return
    def decrease_active_ch_counter(self): # ΜΕΙΩΝΕΙ ΤΟΝ ΑΡΙΘΜΟ ΤΩΝ ΠΡΟΚΛΗΣΕΩΝ ΠΟΥ ΕΧΕΙ ΔΩΣΕΙ Η ΔΕΚΤΕΙ Ο ΠΑΙΚΤΗΣ
        self.active_ch_counter = self.active_ch_counter - 1
        return
    def check_active_challenges(self): # ΕΛΕΓΧΕΙ ΑΝ Ο ΑΡΙΘΜΟΣ ΤΩΝ ΠΡΟΚΛΗΣΕΩΝ ΠΟΥ ΕΧΟΥΝΕ ΔΟΘΕΙ ΞΕΠΕΡΝΑ ΤΟ ΜΑΧΙΜΟΥΜ ΚΑΙ ΑΛΛΑΖΕΙ ΤΗΝ ΚΑΤΑΣΤΑΣΗ 
        if(self.active_ch_counter<MAX_ACTIVE_CHALLENGES):
            self.active_ch = True
        else: self.active_ch =False
        return
 ################ΟΡΙΣΜΟΣ ΚΛΑΣΗ ΜΑΤΣ ΠΡΟΚΛΗΣΗΣ###################
class Challenge_Match:
    def __init__(self,t_rank,g_rank,t_name,g_name,t_surname,g_surname,date,days,challenge_wildcard):
        self.t_rank = t_rank    #ΚΑΤΑΤΑΞΗ ΠΑΙΚΤΗ ΣΤΟΧΟΥ
        self.g_rank = g_rank    #ΚΑΤΑΤΑΞΗ ΠΑΙΚΤΗ ΠΟΥ ΠΡΟΚΑΛΕΣΕ
        self.t_name = t_name    # ΟΝΟΜΑ ΣΤΟΧΟΥ
        self.g_name = g_name    # ΟΝΟΜΑ ΠΑΙΚΤΗ ΠΟΥ ΠΡΟΚΑΛΕΣΕ
        self.t_surname = t_surname  # ΕΠΙΘΕΤΟ ΠΑΙΚΤΗ ΣΤΟΧΟΥ
        self.g_surname = g_surname  # ΕΠΙΘΕΤΟ ΠΑΙΚΤΗ ΠΟΥ ΠΡΟΚΑΛΕΣΕ
        self.date = date # O MHNAΣ ΠΟΥ ΚΛΕΙΣΤΗΚΕ ΤΟ ΜΑΤΣ#
        self.days = days # ΟΙ ΗΜΕΡΕΣ ΠΟΥ ΜΕΝΟΥΝΕ ΩΣΤΕ ΤΟ ΜΑΤΣ ΝΑ ΓΙΝΕΙ ΑΠΟΔΕΚΤΟ
        self.challenge_wildcard = challenge_wildcard # Η μεταβλητη αυτη δηλωνει ενα ματσ που εγινε τσαλεντζ με χρηση wildcard
        return
##################ΟΡΙΣΜΟΣ ΚΛΑΣΗΣ ΙΣΤΟΡΙΚΟΥ ΜΑΤΣ ΠΡΟΚΛΗΣΕΩΝ#####################
class Statistics:
    def __init__(self, t_rank, t_name, t_surname,t_sets, g_rank,g_name,g_surname,g_sets,date):
        self.t_rank = t_rank
        self.t_name = t_name
        self.t_surname = t_surname
        self.t_sets = t_sets
        self.g_rank = g_rank
        self.g_name = g_name
        self.g_surname = g_surname
        self.g_sets = g_sets
        self.date = date
        return
##################################################################################
def make_statistics(list_statistics,t_rank,t_name,t_surname,t_sets,g_rank,g_name,g_surname,g_sets,date): #ΣΥΝΑΡΤΗΣΗ ΠΟΥ ΔΗΜΙΟΥΡΓΕΙ ΤΟ ΙΣΤΟΡΙΚΟ ΤΩΝ ΠΑΙΚΤΩΝ ΣΕ ΛΙΣΤΑ
    list_statistics.append(Statistics(t_rank, t_name, t_surname, t_sets, g_rank, g_name,g_surname,g_sets, date))
    return      
def change_wildcard_after_reject(serial,list_matches,list_players): # ΣΥΝΑΡΤΗΣΗ ΠΟΥ ΕΠΑΝΑΦΕΡΕΙ ΤΟ WILDCARD ΜΕΤΑ ΑΠΟ ΑΡΝΗΣΗ ΜΑΤΣ ΓΙΑ ΤΟΝ ΠΑΙΚΤΗ
        name = list_matches[serial].g_name
        surname = list_matches[serial].g_surname
        for i in range(0,len(list_players)):
            if(list_players[i].name == name and list_players[i].surname == surname):
                list_players[i].wildcard = True
        return
def make_player(list_a,name,surname,age): #ΔΗΜΙΟΥΡΓΕΙ ΤΟ ΑΝΤΙΚΕΙΜΕΝΟ ΤΗΣ ΚΛΑΣΗΣ ΠΑΙΚΤΗ ΚΑΙ ΤΟ ΑΠΟΘΗΚΕΥΕΙ ΣΕ ΛΙΣΤΑ
    list_a.append(Player(name,surname,age,m_pl=0,m_won=0,m_lost=0,sets_pl=0,sets_won=0,sets_lost=0,wildcard=True,active_ch_counter=0,active_ch=True))
    return  
def make_challenge(list_a,t_rank,g_rank,t_name,g_name,t_surname,g_surname,date,days,challenge_wildcard):  #ΔΗΜΙΟΥΡΓΕΙ ΤΟ ΑΝΤΙΚΕΙΜΕΝΟ ΤΗΣ ΚΛΑΣΗΣ ΜΑΤΣ ΠΡΟΚΛΗΣΗΣ ΚΑΙ ΤΟ ΑΠΟΘΗΚΕΥΕΙ ΣΕ ΛΙΣΤΑ
    list_a.append(Challenge_Match(t_rank,g_rank,t_name,g_name,t_surname,g_surname,date,days,challenge_wildcard))
    return
def delete_object(list_a,index): #ΣΒΗΝΕΙ ΕΝΑ ΑΝΤΙΚΕΙΜΕΝΟ ΑΠΟ ΛΙΣΤΑ
    del list_a[index]
    return
def delete_ch_match(list_a,ranking): #ΣΒΗΝΕΙ ΕΝΑ ΑΝΤΙΚΕΙΜΕΝΟ ΑΠΟ ΤΗΝ ΛΙΣΤΑ 
    for i in range(len(list_a)):
        if(list_a[i].t_rank == ranking or list_a[i].g_rank == ranking):
            del list_a[i]
        return

def swap(list_a,winner_position,loser_position):# Η ΣΥΝΑΡΤΗΣΗ ΑΥΤΗ ΔΕΧΕΤΕ ΜΙΑ ΛΙΣΤΑ ΚΑΙ ΔΥΟ ΑΡΙΘΜΟΥΣ ΚΑΤΑΤΑΞΗΣ ΚΑΙ ΕΠΙΣΤΡΕΦΕΙ ΤΟΝ ΛΙΣΤΑ ΜΕ ΒΑΣΗ ΤΗΝ ΑΛΛΑΓΗ ΣΤΗΝ ΚΑΤΑΤΑΞΗ 
   if(winner_position <= loser_position):
       return
   else:
    temp_winner = list_a[winner_position]
    tmp = list_a[winner_position-1]
    list_a[winner_position-1] = temp_winner
    list_a[winner_position] = tmp
    return swap(list_a,winner_position-1,loser_position)
    
def clock_is_ticking(list_matches,list_players): # Συναρτηση που ελεγχει αν περασε ενα χρονικο διαστημα απο την εκδοση ενος ματς και αν εχει γινει αποδεκτο η αν ενημερωθηκε,αν οχι το σβηνει με την παροδο του χρονικου διαστηματος,στην ουσια απλα συγκρινει ημερομηνιες και χρονο
    present= datetime.now().replace(second=0,microsecond=0) #+timedelta(days=2) #αλλαζοντας τιμη στο timedelta μπορει κανεις να τεσταρει αν σβηνει τα ματς η συναρτηση αυτη με την παροδο 15 ημερων
    tmp = []
    for i in range(0,len(list_matches)):
        match_date = list_matches[i].date
        if((present-match_date)> timedelta(days=15)):
            tmp.append(i)
        if((present-list_matches[i].date) < timedelta(days=15)):
            list_matches[i].days = timedelta(days=15) - (present-match_date)
    for i in range(0,len(tmp)):
        j = tmp[i]
        if(list_matches[j].challenge_wildcard==True):
            list_players[list_matches[j].g_rank].wildcard = True
            list_players[list_matches[j].g_rank].decrease_active_ch_counter()
        if(list_matches==ch_valid):
            list_players[list_matches[j].t_rank].decrease_active_ch_counter()
            list_players[list_matches[j].t_rank].check_active_challenges()
            
        delete_object(list_matches,j)    
    return

###################ΛΟΓΙΚΗ ΠΡΟΚΛΗΣΕΩΝ##############################################
###TRUE ΕΓΚΥΡΗ ΠΡΟΚΛΗΣΗ                                                          #
###FALSE ΜΗ ΕΓΚΥΡΗ                                                               #
### Η ΚΥΡΙΑ ΛΟΓΙΚΗ ΓΙΑ ΤΟ ΕΑΝ ΜΠΟΡΕΙ ΑΝ ΔΟΘΕΙ ΜΙΑ ΠΡΟΚΛΗΣΗ                       #
##################################################################################

def main_logic(t_name,t_surname,t_rank,g_name,g_surname,g_rank,list_a,list_b): # ΑΥΤΗ Η ΣΥΝΑΡΤΗΣΗ ΕΙΝΑΙ Η ΚΥΡΙΑ ΛΟΓΙΚΗ ΤΩΝ ΝΕΩΝ ΠΡΟΚΛΗΣΕΩΝ
    if(len(list_b)==0):
        if( g_rank>t_rank and abs(g_rank-t_rank)<=MAX_RANKING_CHALLENGE and list_a[g_rank].wildcard == False and list_a[g_rank].active_ch == True and list_a[t_rank].active_ch == True):
            return True
        if(g_rank>t_rank and list_a[g_rank].wildcard == True and list_a[g_rank].active_ch == True and list_a[t_rank].active_ch == True):
            return True
    else:
        for i in range(0,len(list_b)):
            if(list_b[i].t_name!= t_name and list_b[i].t_surname != t_surname and list_b[i].g_name != g_name and list_b[i].g_surname != g_surname or list_b[i].t_name!= g_name and list_b[i].t_surname != g_surname and list_b[i].g_name != t_name and list_b[i].g_surname != t_surname):
                if( g_rank>t_rank and abs(g_rank-t_rank)<=MAX_RANKING_CHALLENGE and list_a[g_rank].wildcard == False and list_a[g_rank].active_ch == True and list_a[t_rank].active_ch == True):
                    return True
                if(g_rank>t_rank  and list_a[g_rank].wildcard == True and list_a[g_rank].active_ch == True and list_a[t_rank].active_ch == True):
                    return True
    return False


####################################################################
### ΜΕΝΟΥ###########################################################
####################################################################
###ΑΥΤΗ Η ΚΛΑΣΗ ΕΙΝΑΙ ΤΟ ΚΥΡΙΟ ΜΕΝΟΥ###
class MAIN(tkinter.Tk):
    def __init__(self):
        super().__init__()
        self.title("Πρόγραμμα Κατάταξης Τέννις")
        self.screen_width = self.winfo_screenwidth()
        self.screen_height = self.winfo_screenheight()
        self.geometry(f"{self.screen_width}x{self.screen_height}")
        #self.attributes('-fullscreen',True)
        self.configure(background='silver')
        self.treeview = tkinter.ttk.Treeview(selectmode='browse',show='headings',height = 40)
        self.treeview.pack(anchor=tkinter.N,side=tkinter.TOP)
        self.terminate = tkinter.Button(self,height=1,width=6,text = "Εξοδος",activebackground="white",activeforeground="red",bd=8,relief=tkinter.RAISED,bg="red",fg="black",font=("System"),highlightcolor="black",command=self.terminate)
        self.terminate.pack(anchor = tkinter.S,side=tkinter.RIGHT)
        self.settings_button = tkinter.Button(self,height=1,width=9,text = "Ρυθμίσεις",activebackground="white",activeforeground="red",bd=8,relief=tkinter.RAISED,fg="black",bg="green",font=("System"),highlightcolor="black",command=self.settings)
        self.settings_button.pack(anchor = tkinter.S,side=tkinter.RIGHT)
        self.print_ranking_button = tkinter.Button(self,height=1,width=18,text ="Εκτύπωση Κατάταξης",activebackground="white",activeforeground="red",bd=8,relief=tkinter.RAISED,bg="silver",fg="black",font=("System"),highlightcolor="black",command=self.print_ranking)
        self.print_ranking_button.pack(anchor = tkinter.S,side=tkinter.LEFT)
        self.print_stats = tkinter.Button(self,height=1,width=19,text ="Εκτύπωση Στατιστικά",activebackground="white",activeforeground="red",bd=8,relief=tkinter.RAISED,bg="silver",fg="black",font=("System"),highlightcolor="black",command=self.print_ranking_stats)
        self.print_stats.pack(anchor = tkinter.S,side=tkinter.LEFT)
        self.print_stats = tkinter.Button(self,height=1,width=19,text ="Εκτύπωση Ιστορικού",activebackground="white",activeforeground="red",bd=8,relief=tkinter.RAISED,bg="silver",fg="black",font=("System"),highlightcolor="black",command=self.print_history)
        self.print_stats.pack(anchor = tkinter.S,side=tkinter.LEFT)
        self.print_out = tkinter.Button(self,height=1,width=17,text ="Εκτύπωση Εκρρεμών",activebackground="white",activeforeground="red",bd=8,relief=tkinter.RAISED,bg="silver",fg="black",font=("System"),highlightcolor="black",command=self.print_out_ch_matches)
        self.print_out.pack(anchor = tkinter.S,side=tkinter.LEFT)
        self.print_valid = tkinter.Button(self,height=1,width=16,text ="Εκτύπωση Ενεργών",activebackground="white",activeforeground="red",bd=8,relief=tkinter.RAISED,bg="silver",fg="black",font=("System"),highlightcolor="black",command=self.print_valid_ch_matches)
        self.print_valid.pack(anchor = tkinter.S,side=tkinter.LEFT)
        self.man_valid = tkinter.Button(self,height=1,width=18,text ="Διαχείριση Ενεργών",activebackground="white",activeforeground="red",bd=8,relief=tkinter.RAISED,bg="yellow",fg="black",font=("System"),highlightcolor="black",command=self.ch_valid_man)
        self.man_valid.pack(anchor = tkinter.S,side=tkinter.LEFT)
        self.man_out = tkinter.Button(self,height=1,width=18,text ="Διαχείριση Εκρεμών",activebackground="white",activeforeground="red",bd=8,relief=tkinter.RAISED,bg="yellow",fg="black",font=("System"),highlightcolor="black",command=self.ch_out_man)
        self.man_out.pack(anchor = tkinter.S,side=tkinter.LEFT)
        self.ch_new_match = tkinter.Button(self,height=1,width=18,text ="Νέο Μάτς Πρόκλησης",activebackground="white",activeforeground="red",bd=8,relief=tkinter.RAISED,bg="yellow",fg="black",font=("System"),highlightcolor="black",command=self.new_ch_match)
        self.ch_new_match.pack(anchor = tkinter.S,side=tkinter.LEFT)
        self.del_pl = tkinter.Button(self,height=1,width=15,text ="Διαγραφή Παίχτη",activebackground="white",activeforeground="red",bd=8,relief=tkinter.RAISED,bg="cyan",fg="black",font=("System"),highlightcolor="black",command=self.del_player)
        self.del_pl.pack(anchor = tkinter.S,side=tkinter.LEFT)
        self.new_pl = tkinter.Button(self,height=1,width=19,text ="Εγγραφή Νέου Παίχτη",activebackground="white",activeforeground="red",bd=8,relief=tkinter.RAISED,bg="cyan",fg="black",font=("System"),highlightcolor="black",command=self.add_player)
        self.new_pl.pack(anchor = tkinter.S,side=tkinter.LEFT)
   
    ####ΚΟΥΜΠΙ ΤΕΡΜΑΤΙΣΜΟΥ####
    def terminate(self):
        global filename_player_data
        global filename_out_ch
        global filename_valid_ch
        clock_is_ticking(ch_out,pl_data)
        clock_is_ticking(ch_valid,pl_data)
        write_excel(filename_ranking,pl_data,ch_out,ch_valid,statistics)
        root.destroy()
    ###Τυπωνει το ιστορικο###
    def print_history_stats(self,name,surname):
        for item in self.treeview.get_children():
            self.treeview.delete(item)   
        self.treeview['columns'] = ("#1","#2","#3","#4","#5","#6","#7","#8","#9")
        self.treeview.heading("#1",text="Κατάταξη")
        self.treeview.heading("#2",text="Oνομα")
        self.treeview.heading("#3",text="Επίθετο")
        self.treeview.heading("#4",text="Σετς")
        self.treeview.heading("#5",text="Κατάταξη")
        self.treeview.heading("#6",text="Ονομα")
        self.treeview.heading("#7",text="Eπίθετο")
        self.treeview.heading("#8",text="Σετς")
        self.treeview.heading("#9",text="Ημερομηνία")
        for index in range(0,len(statistics)):
            if(statistics[index].g_name == name and statistics[index].g_surname == surname or statistics[index].t_name == name and statistics[index].t_surname == surname):
                self.treeview.insert('',tkinter.END,values=(statistics[index].t_rank+1,statistics[index].t_name,statistics[index].t_surname,statistics[index].t_sets,statistics[index].g_rank+1,statistics[index].g_name,statistics[index].g_surname,statistics[index].g_sets,statistics[index].date)) 
        return                                       
    ###ΤΥΠΩΝΕΙ ΤΗΝ ΒΑΣΙΚΗ ΚΑΤΑΤΑΞΗ###
    def print_ranking(self):
        for item in self.treeview.get_children():
            self.treeview.delete(item)
        self.treeview['columns'] = ("#1","#2","#3","#4","#5","#6")
        self.treeview.heading("#1", text="Κατάταξη")
        self.treeview.heading("#2",text="Ονομα")
        self.treeview.heading("#3",text="Επίθετο")
        self.treeview.heading("#4",text="Ηλικία")
        self.treeview.heading("#5",text="Wildcard")
        self.treeview.heading("#6",text="Ενεργά Παιχνίδια")
        for index in range(0,len(pl_data)):
            self.treeview.insert('',tkinter.END,values=(index+1,pl_data[index].name,pl_data[index].surname,pl_data[index].age,pl_data[index].wildcard,pl_data[index].active_ch))
        return
    
    ###TYΠΩΝΕΙ ΤΗΝ ΚΑΤΑΤΑΞΗ ΜΕ ΣΤΑΤΙΣΤΙΚΑ###
    def print_ranking_stats(self):
        for item in self.treeview.get_children():
            self.treeview.delete(item)
        self.treeview['columns'] = ("#1","#2","#3","#4","#5","#6","#7","#8","#9")
        self.treeview.heading("#1", text="Κατάταξη")
        self.treeview.heading("#2", text="Ονομα")
        self.treeview.heading("#3", text="Επώνυμο")
        self.treeview.heading("#4", text="Παίχτηκαν")
        self.treeview.heading("#5", text="Νίκες")
        self.treeview.heading("#6", text="Ηττες")
        self.treeview.heading("#7", text="Σύνολο Set")
        self.treeview.heading("#8", text="Νικηφόρα Set")
        self.treeview.heading("#9", text="Χαμένα Set")
        for index in range(0,len(pl_data)):
            self.treeview.insert('',tkinter.END,values=(index+1,pl_data[index].name,pl_data[index].surname,pl_data[index].m_pl,pl_data[index].m_won,pl_data[index].m_lost,pl_data[index].sets_pl,pl_data[index].sets_won,pl_data[index].sets_lost))
        return

    ###ΤΥΠΩΝΕΙ ΤΑ ΜΑΤΣ ΠΡΟΚΛΗΣΗΣ ΠΟΥ ΕΙΝΑΙ ΕΚΡΕΜΗ###
    def print_out_ch_matches(self):
        for item in self.treeview.get_children():
            self.treeview.delete(item)
        self.treeview['columns'] = ("#1","#2","#3","#4","#5","#6","#7","#8","#9")
        self.treeview.heading("#1", text = "A/A Πρόκλησης")
        self.treeview.heading("#2", text = "Κατάταξη")
        self.treeview.heading("#3", text = "Ονομα")
        self.treeview.heading("#4", text = "Επώνυμο")
        self.treeview.heading("#5", text = "Κατάταξη")
        self.treeview.heading("#6", text = "Ονομα")
        self.treeview.heading("#7", text = "Επώνυμο")
        self.treeview.heading("#8", text = "Ημερομηνια")
        self.treeview.heading("#9", text = "Υπόλοιπo")
        if(len(ch_out)>0):
            for index in range(0,len(ch_out)):
                self.treeview.insert('',tkinter.END,values=(index+1,ch_out[index].g_rank+1,ch_out[index].g_name,ch_out[index].g_surname,ch_out[index].t_rank+1,ch_out[index].t_name,ch_out[index].t_surname,ch_out[index].date,ch_out[index].days))
        return

    ###ΤΥΠΩΝΕΙ ΤΑ ΜΑΤΣ ΠΡΟΚΛΗΣΗΣ ΠΟΥ ΕΙΝΑΙ ΕΝΕΡΓΑ###
    def print_valid_ch_matches(self):
        for item in self.treeview.get_children():
            self.treeview.delete(item)
        self.treeview['columns'] = ("#1","#2","#3","#4","#5","#6","#7","#8","#9")
        self.treeview.heading("#1", text = "A/A Πρόκλησης")
        self.treeview.heading("#2", text = "Κατάταξη")
        self.treeview.heading("#3", text = "Ονομα")
        self.treeview.heading("#4", text = "Επώνυμο")
        self.treeview.heading("#5", text = "Κατάταξη")
        self.treeview.heading("#6", text = "Ονομα")
        self.treeview.heading("#7", text = "Επώνυμο")
        self.treeview.heading("#8", text = "Ημερομηνια")
        self.treeview.heading("#9", text = "Υπόλοιπο")
        if (len(ch_valid)>0):
            for index in range(0,len(ch_valid)):
                self.treeview.insert('',tkinter.END,values=(index+1,ch_valid[index].t_rank+1,ch_valid[index].t_name,ch_valid[index].t_surname,ch_valid[index].g_rank+1,ch_valid[index].g_name,ch_valid[index].g_surname,ch_valid[index].date,ch_valid[index].days))
        return
    ###ΔΙΑΧΕΙΡΙΣΗ ΕΚΚΡΕΜΩΝ ΜΑΤΣ ΠΡΟΚΛΗΣΗΣ###
    def ch_out_man(self):
        top = OUTSTANDING_CHALLENGE_MATCH(self)
        top.grab_set()
    ###ΕΝΗΜΕΡΩΣΗ ΕΝΕΡΓΩΝ ΜΑΤΣ ΠΡΟΚΛΗΣΗΣ###
    def ch_valid_man(self):
        top = BRIEF_CHALLENGE_MATCH(self)
        top.grab_set()
    ###ΝΕΟ ΜΑΤΣ ΠΡΟΚΛΗΣΗΣ###
    def new_ch_match(self):
        top = NEW_CHALLENGE_MATCH(self)
        top.grab_set()
    ###ΔΙΑΓΡΑΦΗ ΠΑΙΚΤΗ###
    def del_player(self):
        top = DEL_PLAYER(self)
        top.grab_set()
    ###ΕΓΡΑΦΗ ΝΕΟΥ ΠΑΙΧΤΗ###
    def add_player(self):
        top = ADD_PLAYER(self)
        top.grab_set()
    ###ΡΥΘΜΙΣΕΙΣ###
    def settings(self):
        top = SETTINGS(self)
        top.grab_set()
    ####HISTORY####
    def print_history(self):
        top = HISTORY(self)
        top.grab_set()

#####ΚΛΑΣΗ ΥΠΟ ΜΕΝΟΥ ΕΚΤΥΠΩΣΗ ΙΣΤΟΡΙΚΟΥ###
class HISTORY(tkinter.Toplevel):
    def __init__(self,parent):
        super().__init__(parent)
        self.title("Πρόγραμμα Κατάταξης Τέννις")
        self.screen_width = self.winfo_screenwidth()
        self.screen_height = self.winfo_screenheight()
        self.configure(background='silver')
        self.geometry(f"{self.screen_width//4}x{self.screen_height//4}+{self.screen_width//3}+{self.screen_height//3}")
        self.rank = tkinter.IntVar(value=1)
        tkinter.Label(self,bg="silver",text="Αριθμός κατάταξης Παίχτη προς Εκτύπωση Ιστορικού",fg="green",font=("System",12)).pack()
        tkinter.Entry(self,fg="black",bg="white",font=("System",12,"bold"),border=4,textvariable=self.rank).pack(anchor=tkinter.N,side=tkinter.TOP)
        tkinter.Button(self,text='Επιστροφή',activebackground="white",activeforeground="red",fg="black",font=("System",10),highlightcolor="black",bg="silver",padx=10,pady=10,command=self.destroy).pack(anchor=tkinter.S,side=tkinter.LEFT)
        tkinter.Button(self,height=1,width=8,text = "Εκτύπωση",activebackground="white",bg="green",activeforeground="red",bd=8,relief=tkinter.RAISED,fg="black",font=("System",12),highlightcolor="black",command=self.print_history_fuction).pack(anchor=tkinter.S,side=tkinter.RIGHT)
       #####ΚΟΥΜΠΙ ΑΠΟΔΟΧΗΣ ΙΣΤΟΡΙΚΟΥ####
    def print_history_fuction(self):
        try:
            rank = self.rank.get() - 1
        except:
            print("EXCEPT")
            top = FAILED(self)
            top.grab_set()
        if(rank>=0 and rank<=len(pl_data)):
                history_name = pl_data[rank].name
                history_surname = pl_data[rank].surname
                root.print_history_stats(history_name,history_surname)
                top = CONFIRMED(self)
                top.grab_set()
        else:
                print("IF FAILED")
                top = FAILED(self)
                top.grab_set()
        return
###ΚΛΑΣΗ ΥΠΟ ΜΕΝΟΥ ΑΠΟΔΟΧΗ ΕΙΣΟΔΟΥ###

class CONFIRMED(tkinter.Toplevel): 
    def __init__(self,parent): 
        super().__init__(parent)
        self.title("Πρόγραμμα Κατάταξης Τέννις")
        self.screen_width = self.winfo_screenwidth()
        self.screen_height = self.winfo_screenheight()
        self.configure(background='silver')
        self.geometry(f"{self.screen_width//8}x{self.screen_height//8}+{self.screen_width//3}+{self.screen_height//3}")
        tkinter.Label(self,bg="silver",text="Αποδοχη",fg="green",font=("System",12)).pack()
        tkinter.Button(self,text='Επιστροφή',activebackground="white",activeforeground="red",fg="black",font=("System",10),highlightcolor="black",bg="silver",padx=10,pady=10,command=self.destroy).pack(anchor=tkinter.S,side=tkinter.BOTTOM)

###ΚΛΑΣΗ ΥΠΟ ΜΕΝΟΥ ΑΠΟΡΡΙΨΗ ΕΙΣΟΔΟΥ###
class FAILED(tkinter.Toplevel): 
    def __init__(self,parent): 
        super().__init__(parent)
        self.title("Πρόγραμμα Κατάταξης Τέννις")
        self.screen_width = self.winfo_screenwidth()
        self.screen_height = self.winfo_screenheight()
        self.configure(background='silver')
        self.geometry(f"{self.screen_width//8}x{self.screen_height//8}+{self.screen_width//3}+{self.screen_height//3}")
        tkinter.Label(self,bg="silver",text="Απόρριψη",fg="red",font=("System",12)).pack()
        tkinter.Button(self,text='Επιστροφή',activebackground="white",activeforeground="red",bg="silver",fg="black",font=("System",10),highlightcolor="black",padx=10,pady=10,command=self.destroy).pack(anchor=tkinter.S,side=tkinter.BOTTOM)
###ΚΛΑΣΗ ΥΠΟ ΜΕΝΟΥ ΕΓΓΡΑΦΗ ΝΕΟΥ ΠΑΙΚΤΗ###
class ADD_PLAYER(tkinter.Toplevel):
    def __init__(self,parent): 
        super().__init__(parent)
        self.title("Πρόγραμμα Κατάταξης Τέννις")
        self.screen_width = self.winfo_screenwidth()
        self.screen_height = self.winfo_screenheight()
        self.configure(background='silver')
        self.geometry(f"{self.screen_width//4}x{self.screen_height//4}+{self.screen_width//3}+{self.screen_height//3}")
        self.name = tkinter.StringVar()
        self.surname = tkinter.StringVar()
        self.age = tkinter.IntVar(value=16)
        tkinter.Label(self,text="Ονομα",bg="silver",fg="black",font=("System",12,"bold")).pack(anchor=tkinter.N,side=tkinter.TOP)
        tkinter.Entry(self,fg="black",bg="white",font=("System",12,"bold"),border=4,textvariable=self.name).pack(anchor=tkinter.N,side=tkinter.TOP)
        tkinter.Label(self,text="Επώνυμο",bg="silver",fg="black",font=("System",12,"bold")).pack(anchor=tkinter.N,side=tkinter.TOP)
        tkinter.Entry(self,fg="black",bg="white",font=("System",12,"bold"),border=4,textvariable=self.surname).pack(anchor=tkinter.N,side=tkinter.TOP)
        tkinter.Label(self,text="Ηλικία",bg="silver",fg="black",font=("System",12,"bold")).pack(anchor=tkinter.N,side=tkinter.TOP)
        tkinter.Entry(self,fg="black",bg="white",font=("System",12,"bold"),border=4,textvariable=self.age,relief=tkinter.SUNKEN).pack(anchor=tkinter.N,side=tkinter.TOP)
        tkinter.Button(self,height=1,width=30,text = "Επιστροφή",activebackground="white",bg="silver",activeforeground="red",bd=8,relief=tkinter.RAISED,fg="black",font=("System",12),highlightcolor="black",command=self.destroy).pack(anchor=tkinter.SW,side=tkinter.LEFT)
        tkinter.Button(self,height=1,width=30,text = "Εγγραφή",activebackground="white",bg="green",activeforeground="red",bd=8,relief=tkinter.RAISED,fg="black",font=("System",12),highlightcolor="black",command=self.confirm).pack(anchor=tkinter.SW,side=tkinter.RIGHT)
###ΕΠΙΒΕΒΑΙΩΣΗ ΠΡΟΣΘΗΚΗΣ ΠΑΙΚΤΗ###
    def confirm(self):
        try:
                name=self.name.get()
                surname=self.surname.get()
                age=self.age.get()
        except:
                top = FAILED(self)
                top.grab_set()
                
        if(len(name)>=1 and len(surname)>=1 and age>=16):
            make_player(pl_data,name,surname,age) #ΠΡΟΣΘΕΤΕΙ ΤΟΝ ΠΑΙΧΤΗ ΕΑN 16+
            root.print_ranking()
            top = CONFIRMED(self)
            top.grab_set()
        
        

###ΚΛΑΣΗ ΥΠΟ ΜΕΝΟΥ ΔΙΑΓΡΑΦΗ ΠΑΙΚΤΗ####
class DEL_PLAYER(tkinter.Toplevel):
    def __init__(self,parent): 
        super().__init__(parent)
        self.title("Πρόγραμμα Κατάταξης Τέννις")
        self.screen_width = self.winfo_screenwidth()
        self.screen_height = self.winfo_screenheight()
        self.configure(background='silver')
        self.geometry(f"{self.screen_width//4}x{self.screen_height//4}+{self.screen_width//3}+{self.screen_height//3}")
        self.variable_del = tkinter.IntVar()
        tkinter.Label(self,text="Αριθμός Κατάταξης",bg="silver",fg="black",font=("System",12,"bold")).pack(anchor=tkinter.N,side=tkinter.TOP)
        tkinter.Entry(self,fg="black",bg="white",font=("System",12,"bold"),border=4,textvariable=self.variable_del).pack(anchor=tkinter.N,side=tkinter.TOP)
        tkinter.Button(self,height=1,width=30,bg="silver",text = "Επιστροφή",activebackground="white",activeforeground="red",bd=8,relief=tkinter.RAISED,fg="black",font=("System",12),highlightcolor="black",command=self.destroy).pack(anchor=tkinter.SW,side=tkinter.LEFT)
        tkinter.Button(self,height=1,width=30,bg="red",text = "Διαγραφή",activebackground="white",activeforeground="red",bd=8,relief=tkinter.RAISED,fg="black",font=("System",12),highlightcolor="black",command=self.confirm_del).pack(anchor=tkinter.SW,side=tkinter.RIGHT)
###EΠΙΒΕΒΑΙΩΣΗ ΔΙΑΓΡΑΦΗ ΠΑΙΚΤΗ###
    def confirm_del(self):
        try:
            var_del_player = self.variable_del.get()-1
        except:
            top = FAILED(self)
            top.grab_set()
               
        if(var_del_player >=0 and var_del_player<=(len(pl_data))):         
            delete_ch_match(ch_out,var_del_player)
            delete_ch_match(ch_valid,var_del_player)
            delete_object(pl_data,var_del_player)
            root.print_ranking()
            top = CONFIRMED(self)
            top.grab_set()
        else:
            top = FAILED(self)
            top.grab_set()
               
###ΚΛΑΣΗ ΥΠΟΜΕΝΟΥ ΡΥΘΜΙΣΕΙΣ ##
class SETTINGS(tkinter.Toplevel):
    def __init__(self,parent):
        super().__init__(parent)
        self.title("Πρόγραμμα Κατάταξης Τέννις")
        self.screen_width = self.winfo_screenwidth()
        self.screen_height = self.winfo_screenheight()
        self.configure(background='silver')
        self.geometry(f"{self.screen_width//4}x{self.screen_height//4}+{self.screen_width//3}+{self.screen_height//3}")
        self.wildcard = tkinter.IntVar()
        self.max_ch_ranking = tkinter.IntVar()
        self.max_active = tkinter.IntVar() 
        self.max_sets = tkinter.IntVar() 
        tkinter.Label(self,bg="silver",text="Μάτς που παίζονται για Μπαλαντέρ",fg="black",font=("System",12,"bold")).pack(anchor=tkinter.N,side=tkinter.TOP)
        tkinter.Entry(self,bg="white",fg="black",font=("System",12,"bold"),border=4,textvariable=self.wildcard).pack(anchor=tkinter.N,side=tkinter.TOP)
        tkinter.Label(self,bg="silver",text="Απόσταση στην Κατάταξη μεταξύ παικτών για προκλήση",fg="black",font=("System",12,"bold")).pack(anchor=tkinter.N,side=tkinter.TOP)
        tkinter.Entry(self,bg="white",fg="black",font=("System",12,"bold"),border=4,textvariable=self.max_ch_ranking).pack(anchor=tkinter.N,side=tkinter.TOP)
        tkinter.Label(self,bg="silver",text="Aριθμός σετ που πρέπει να κερδίσει ο παίκτης για να πάρει το παιχνίδι",fg="black",font=("System",12,"bold")).pack(anchor=tkinter.N,side=tkinter.TOP)
        tkinter.Entry(self,bg="white",fg="black",font=("System",12,"bold"),border=4,textvariable=self.max_sets).pack(anchor=tkinter.N,side=tkinter.TOP)
        tkinter.Button(self,height=1,width=10,bg="silver",text = "Επιστροφή",activebackground="white",activeforeground="red",bd=8,relief=tkinter.RAISED,fg="black",font=("System",12),highlightcolor="black",command=self.destroy).pack(anchor=tkinter.SW,side=tkinter.LEFT)
        tkinter.Button(self,height=1,width=10,bg="green",text = "Εγγραφή",activebackground="white",activeforeground="red",bd=8,relief=tkinter.RAISED,fg="black",font=("System",12),highlightcolor="black",command=self.define_settings).pack(anchor=tkinter.SW,side=tkinter.RIGHT)
### ΡΥΘΜΙΣΕΙΣ###
    def define_settings(self):
        global WILDCARD
        global MAX_RANKING_CHALLENGE
        global POINT_SET_MATCH
        try:
            WILDCARD = self.wildcard.get()
            MAX_RANKING_CHALLENGE = self.max_ch_ranking.get()
            POINT_SET_MATCH = self.max_sets.get()
            write_config()
            top = CONFIRMED(self)
            top.grab_set()
        except:
            top = FAILED(self)
            top.grab_set()
        return 
       
    
###ΚΛΑΣΗ ΥΠΟΜΕΝΟΥ ΝΕΟ ΜΑΤΣ ΠΡΟΚΛΗΣΗΣ###
class NEW_CHALLENGE_MATCH(tkinter.Toplevel):
    def __init__(self,parent):
        super().__init__(parent)
        self.title("Πρόγραμμα Κατάταξης Τέννις")
        self.screen_width = self.winfo_screenwidth()
        self.screen_height = self.winfo_screenheight()
        self.configure(background='silver')
        self.geometry(f"{self.screen_width//4}x{self.screen_height//4}+{self.screen_width//3}+{self.screen_height//3}")
        self.challenger = tkinter.IntVar(value=1)
        self.champion = tkinter.IntVar(value=1)
        tkinter.Label(self,bg="silver",text="Κατάταξη Παίκτη που προκαλεί",fg="black",font=("System",12,"bold")).pack(anchor=tkinter.N,side=tkinter.TOP)
        tkinter.Entry(self,bg="white",fg="black",font=("System",12,"bold"),border=4,textvariable=self.challenger).pack(anchor=tkinter.N,side=tkinter.TOP)
        tkinter.Label(self,bg="silver",text="Κατάταξη Παίκτη αποδέκτης πρόκλησης",fg="black",font=("System",12,"bold")).pack(anchor=tkinter.N,side=tkinter.TOP)
        tkinter.Entry(self,bg="white",fg="black",font=("System",12,"bold"),border=4,textvariable=self.champion).pack(anchor=tkinter.N,side=tkinter.TOP)
        tkinter.Button(self,bg="silver",height=1,width=10,text = "Επιστροφή",activebackground="white",activeforeground="red",bd=8,relief=tkinter.RAISED,fg="black",font=("System",12),highlightcolor="black",command=self.destroy).pack(anchor=tkinter.SW,side=tkinter.LEFT)
        tkinter.Button(self,bg="green",height=1,width=10,text = "Εγγραφή",activebackground="white",activeforeground="red",bd=8,relief=tkinter.RAISED,fg="black",font=("System",12),highlightcolor="black",command=self.confirm_ch_match).pack(anchor=tkinter.SW,side=tkinter.RIGHT)
    def confirm_ch_match(self):
        try:
            g_rank = self.challenger.get()-1
            t_rank = self.champion.get()-1
           
        except:
            top = FAILED(self)
            top.grab_set()
        
        if( g_rank >=0 and g_rank <= (len(pl_data))  and t_rank >=0 and t_rank <=len(pl_data) and g_rank>t_rank):
            
           ### ΕΛΕΓΧΕΙ ΑΝ ΟΙ ΑΡΙΘΜΟΙ ΕΙΝΑΙ ΣΤΗΝ ΕΠΙΤΡΕΠΟΜΕΝΗ ΑΠΟΣΤΑΣΗ
            if(main_logic(pl_data[t_rank].name,pl_data[t_rank].surname,t_rank,pl_data[g_rank].name,pl_data[g_rank].surname,g_rank,pl_data,ch_out)==True and pl_data[g_rank].active_ch==True and pl_data[t_rank].active_ch == True and pl_data[g_rank].active_ch_counter<MAX_ACTIVE_CHALLENGES and pl_data[t_rank].active_ch_counter<MAX_ACTIVE_CHALLENGES):
                
                pl_data[g_rank].increase_active_ch_counter()
                pl_data[g_rank].check_active_challenges()
                if(abs(g_rank-t_rank)>=MAX_RANKING_CHALLENGE and pl_data[g_rank].wildcard==True):
                    challenge_made = True
                    pl_data[g_rank].change_wild_state()
                else:
                    challenge_made = False
                match_day = datetime.now().replace(second=0,microsecond=0)
                make_challenge(ch_out,t_rank,g_rank,pl_data[t_rank].name,pl_data[g_rank].name,pl_data[t_rank].surname,pl_data[g_rank].surname,match_day,timedelta(days=15),challenge_made)
                clock_is_ticking(ch_out,pl_data)
                root.print_out_ch_matches()
                top = CONFIRMED(self)
                top.grab_set()
                    
                    
            else:
                top = FAILED(self)
                top.grab_set()
                  
        else:
            top = FAILED(self)
            top.grab_set()          

###ΚΛΑΣΗ ΥΠΟΜΕΝΟΥ ΔΙΑΧΕΙΡΙΣΗ ΕΝΕΡΓΩΝ ΜΑΤΣ ΠΡΟΚΛΗΣΗΣ###
class BRIEF_CHALLENGE_MATCH(tkinter.Toplevel):
    def __init__(self,parent):
        super().__init__(parent)
        self.title("Πρόγραμμα Κατάταξης Τέννις")
        self.screen_width = self.winfo_screenwidth()
        self.screen_height = self.winfo_screenheight()
        self.configure(background='silver')
        self.geometry(f"{self.screen_width//3}x{self.screen_height//3}+{self.screen_width//3}+{self.screen_height//3}")
        self.serial = tkinter.IntVar()
        self.winner = tkinter.IntVar()
        self.loser = tkinter.IntVar()
        self.sets_winner = tkinter.IntVar()
        self.sets_loser = tkinter.IntVar()
        tkinter.Label(self,bg="silver",text="Αριθμός μάτς κατάταξης προς ενημέρωση",fg="black",font=("System",12,"bold")).pack(anchor=tkinter.N,side=tkinter.TOP)
        tkinter.Entry(self,bg="white",fg="black",font=("System",12,"bold"),border=4,textvariable=self.serial).pack(anchor=tkinter.N,side=tkinter.TOP)
        tkinter.Label(self,bg="silver",text="Αριθμός κατάταξης παίχτη που νίκησε",fg="black",font=("System",12,"bold")).pack(anchor=tkinter.N,side=tkinter.TOP)
        tkinter.Entry(self,bg="white",fg="black",font=("System",12,"bold"),border=4,textvariable=self.winner).pack(anchor=tkinter.N,side=tkinter.TOP)
        tkinter.Label(self,bg="silver",text="Σετ που πήρε ο Νικητής",fg="black",font=("System",12,"bold")).pack(anchor=tkinter.N,side=tkinter.TOP)
        tkinter.Entry(self,bg="white",fg="black",font=("System",12,"bold"),border=4,textvariable=self.sets_winner).pack(anchor=tkinter.N,side=tkinter.TOP)
        tkinter.Label(self,bg="silver",text="Αριθμός κατάταξης παίχτη που ηττήθηκε",fg="black",font=("System",12,"bold")).pack(anchor=tkinter.N,side=tkinter.TOP)
        tkinter.Entry(self,bg="white",fg="black",font=("System",12,"bold"),border=4,textvariable=self.loser).pack(anchor=tkinter.N,side=tkinter.TOP)
        tkinter.Label(self,bg="silver",text="Σετ που πήρε ο Ηττημένος",fg="black",font=("System",12,"bold")).pack(anchor=tkinter.N,side=tkinter.TOP)
        tkinter.Entry(self,bg="white",fg="black",font=("System",12,"bold"),border=4,textvariable=self.sets_loser).pack(anchor=tkinter.N,side=tkinter.TOP)
        tkinter.Button(self,bg="silver",height=1,width=10,text = "Επιστροφή",activebackground="white",activeforeground="red",bd=8,relief=tkinter.RAISED,fg="black",font=("System",12),highlightcolor="black",command=self.destroy).pack(anchor=tkinter.SW,side=tkinter.LEFT)
        tkinter.Button(self,bg="green",height=1,width=10,text = "Εγγραφή",activebackground="white",activeforeground="red",bd=8,relief=tkinter.RAISED,fg="black",font=("System",12),highlightcolor="black",command=self.brief_match).pack(anchor=tkinter.SW,side=tkinter.RIGHT)
    def brief_match(self):
        
        serial = self.serial.get()-1
        winner = self.winner.get()-1
        loser = self.loser.get()-1
        sets_winner = self.sets_winner.get()
        sets_loser = self.sets_loser.get()
        today = datetime.now().replace(second=0,microsecond=0)
        if( sets_winner>=1 and sets_winner<=POINT_SET_MATCH and sets_loser>=0 and sets_loser<=POINT_SET_MATCH and sets_winner>sets_loser and winner>=0 and winner<=len(pl_data) and loser>=0 and loser<=len(pl_data) and serial>=0 and serial<=(len(ch_valid))):#EΑΝ ΟΙ ΑΡΙΘΜΟΙ ΥΠΟΚΕΙΝΤΑΙ ΣΤΟΥΣ ΚΑΝΟΝΕΣ ΓΙΑ ΤΑ ΠΑΙΧΝΔΙΑ
            pl_data[winner].increase_m_pl()
            pl_data[loser].increase_m_pl()
            pl_data[winner].increase_m_won()
            pl_data[loser].increase_m_lost()
            pl_data[winner].increase_sets_won(sets_winner)
            pl_data[winner].increase_sets_lost(sets_loser)
            pl_data[loser].increase_sets_won(sets_loser)
            pl_data[loser].increase_sets_lost(sets_winner)
            pl_data[winner].increase_sets_pl()
            pl_data[loser].increase_sets_pl()
            pl_data[winner].decrease_active_ch_counter()
            pl_data[loser].decrease_active_ch_counter()
            pl_data[winner].check_active_challenges()
            pl_data[loser].check_active_challenges()
            pl_data[loser].check_wild_state()
            pl_data[winner].check_wild_state()
            if (ch_valid[serial].t_rank == winner):
                make_statistics(statistics,ch_valid[serial].t_rank,ch_valid[serial].t_name,ch_valid[serial].t_surname,sets_winner,ch_valid[serial].g_rank,ch_valid[serial].g_name,ch_valid[serial].g_surname,sets_loser,today)
            if (ch_valid[serial].g_rank == winner):
                make_statistics(statistics,ch_valid[serial].t_rank,ch_valid[serial].t_name,ch_valid[serial].t_surname,sets_loser,ch_valid[serial].g_rank,ch_valid[serial].g_name,ch_valid[serial].g_surname,sets_winner,today)
            swap(pl_data,winner,loser)
            delete_object(ch_valid,serial)
            root.print_valid_ch_matches()
            top = CONFIRMED(self)
            top.grab_set() 
        else:
            top = FAILED(self)
            top.grab_set()
                
###ΚΛΑΣΗ ΥΠΟΜΕΝΟΥ ΔΙΑΧΕΙΡΙΣΗ ΕΚΡΕΜΜΩΝ ΜΑΤΣ ΠΡΟΚΛΗΣΗΣ###
class OUTSTANDING_CHALLENGE_MATCH(tkinter.Toplevel):
    def __init__(self,parent):
        super().__init__(parent)
        self.title("Πρόγραμμα Κατάταξης Τέννις")
        self.screen_width = self.winfo_screenwidth()
        self.screen_height = self.winfo_screenheight()
        self.configure(background='silver')
        self.geometry(f"{self.screen_width//4}x{self.screen_height//4}+{self.screen_width//3}+{self.screen_height//3}")
        self.match = tkinter.IntVar()
        tkinter.Label(self,bg="silver",text="Αριθμός εκρεμούς μάτς κατάταξης",fg="black",font=("System",12,"bold")).pack(anchor=tkinter.N,side=tkinter.TOP)
        tkinter.Entry(self,bg="white",fg="black",font=("System",12,"bold"),border=4,textvariable=self.match).pack(anchor=tkinter.N,side=tkinter.TOP)
        tkinter.Button(self,bg="silver",height=1,width=10,text = "Επιστροφή",activebackground="white",activeforeground="red",bd=8,relief=tkinter.RAISED,fg="black",font=("System",12),highlightcolor="black",command=self.destroy).pack(anchor=tkinter.SW,side=tkinter.LEFT)
        tkinter.Button(self,bg="green",height=1,width=10,text = "Αποδοχή",activebackground="white",activeforeground="red",bd=8,relief=tkinter.RAISED,fg="black",font=("System",12),highlightcolor="black",command=self.confirm_match).pack(anchor=tkinter.SW,side=tkinter.RIGHT)
        tkinter.Button(self,bg="red",height=1,width=10,text = "Απόρριψη",activebackground="white",activeforeground="red",bd=8,relief=tkinter.RAISED,fg="black",font=("System",12),highlightcolor="black",command=self.deny_match).pack(anchor=tkinter.SW,side=tkinter.RIGHT)
    ###ΑΠΟΔΟΧΗ ΕΚΡΕΜΜΟΥΣ ΜΑΤΣ ΠΡΟΚΛΗΣΗΣ###
    def confirm_match(self):
        try:
            number = self.match.get() - 1
        except:
            pop_up = FAILED(self)
            pop_up.grab_set()
        if(number>=0 and number <=(len(ch_out))):
            t_rank = ch_out[number].t_rank
            g_rank = ch_out[number].g_rank
            match_day = datetime.now().replace(second=0,microsecond=0)
            if(pl_data[t_rank].active_ch==True):
                pl_data[t_rank].increase_active_ch_counter()
                pl_data[t_rank].check_active_challenges()
                make_challenge(ch_valid,t_rank,g_rank,pl_data[t_rank].name,pl_data[g_rank].name,pl_data[t_rank].surname,pl_data[g_rank].surname,match_day,timedelta(days=15),ch_out[number].challenge_wildcard)
                clock_is_ticking(ch_valid,pl_data)             
                delete_object(ch_out,number)
                root.print_valid_ch_matches()
                pop_up = CONFIRMED(self)
                pop_up.grab_set()
            else:
                    pop_up = FAILED(self)
                    pop_up.grab_set()
        else:
                pop_up = FAILED(self)
                pop_up.grab_set()

    ###ΑΡΝΗΣΗ ΕΚΡΕΜΟΥΣ ΜΑΤΣ ΠΡΟΚΛΗΣΗΣ###
    def deny_match(self):
        try:
            number = self.match.get() -1
        except:
            pop_up = FAILED(self)
            pop_up.grab_set()
        if(number>=0 and number<=len(ch_out)):
            pl_data[ch_out[number].g_rank].decrease_active_ch_counter()
            pl_data[ch_out[number].g_rank].check_active_challenges()
            pl_data[ch_out[number].t_rank].decrease_active_ch_counter()
            pl_data[ch_out[number].t_rank].check_active_challenges()
            change_wildcard_after_reject(number,ch_out,pl_data)
            delete_object(ch_out,number)
            root.print_out_ch_matches()
            pop_up = CONFIRMED(self)
            pop_up.grab_set()
        else:
            pop_up = FAILED(self)
            pop_up.grab_set()
######################ΣΥΝΑΡΤΗΣΕΙΣ ΠΟΥ ΓΡΑΦΟΥΝΕ ΣΕ ΑΡΧΕΙΑ ΕΞΕΛ ΚΑΙ ΙΝΙ#################            
def open_excel(filename,list_a,list_b,list_c,list_d):## ΑΝΟΙΓΕΙ ΕΝΑ ΕΞΕΛ ΑΡΧΕΙΟ ΚΑΙ ΤΟ ΔΙΑΒΑΖΕΙ ΚΑΙ ΦΤΙΑΧΝΕΙ ΤΙΣ ΛΙΣΤΕΣ ΕΙΣΟΔΟΥ
    try:
        workbook = openpyxl.load_workbook(filename)
        workbook.active = 0
        worksheet = workbook.active
    except:
        workbook = Workbook()
        worksheet = workbook.active
        workbook.save('ranking.xlsx')
    if worksheet is None:
        worksheet = workbook.create_sheet()
    for row in worksheet.iter_rows(): # ΔΙΑΒΑΖΕΙ ΤΑ ΣΤΟΙΧΕΙΑ ΤΩΝ ΠΑΙΚΤΩΝ
        for cell in row:
            if 'A' in cell.coordinate:
                name = cell.value
            if 'B' in cell.coordinate:
                surname = cell.value
            if 'C' in cell.coordinate:
                age = cell.value
            if 'D' in cell.coordinate:    
                m_pl = cell.value
            if 'E' in cell.coordinate:        
                m_won = cell.value
            if 'F' in cell.coordinate:
                m_lost = cell.value
            if 'G' in cell.coordinate:
                sets_pl = cell.value
            if 'H' in cell.coordinate:
                sets_won = cell.value
            if 'I' in cell.coordinate:
                sets_lost = cell.value
            if 'J' in cell.coordinate:
                wildcard = cell.value
            if 'K' in cell.coordinate:
                active_ch_counter = cell.value
            if 'L' in cell.coordinate:
                active_ch = cell.value
                list_a.append(Player(name,surname,age,m_pl,m_won,m_lost,sets_pl,sets_won,sets_lost,wildcard,active_ch_counter,active_ch))
    workbook.active = 1
    worksheet = workbook.active
    if worksheet is None:
        worksheet = workbook.create_sheet()
    for row in worksheet.iter_rows(): # ΔΙΑΒΑΖΕΙ ΤΑ ΣΤΟΙΧΕΙΑ ΤΩΝ ΜΗ ΕΓΚΥΡΩΝ ΜΑΤΣ ΠΡΟΚΛΗΣΗΣ
        for cell in row:
            if 'A' in cell.coordinate:
                t_rank = cell.value
            if 'B' in cell.coordinate:
                g_rank = cell.value
            if 'C' in cell.coordinate:
                t_name = cell.value
            if 'D' in cell.coordinate:
                g_name = cell.value
            if 'E' in cell.coordinate:
                t_surname = cell.value
            if 'F' in cell.coordinate:
                g_surname = cell.value
            if 'G' in cell.coordinate:
                date = cell.value
            if 'H' in cell.coordinate:
                days = cell.value
            if 'I' in cell.coordinate:
                wildcard = cell.value
                list_b.append(Challenge_Match(t_rank,g_rank,t_name,g_name,t_surname,g_surname,date,days,wildcard))
    workbook.active = 2
    worksheet = workbook.active
    if worksheet is None:
        worksheet = workbook.create_sheet() 
    for row in worksheet.iter_rows(): # ΔΙΑΒΑΖΕΙ ΤΑ ΣΤΟΙΧΕΙΑ ΤΩΝ ΕΓΚΥΡΩΝ ΜΑΤΣ ΠΡΟΚΛΗΣΗΣ
        for cell in row:
            if 'A' in cell.coordinate:
                t_rank = cell.value
            if 'B' in cell.coordinate:
                g_rank = cell.value
            if 'C' in cell.coordinate:
                t_name = cell.value
            if 'D' in cell.coordinate:
                g_name = cell.value
            if 'E' in cell.coordinate:
                t_surname = cell.value
            if 'F' in cell.coordinate:
                g_surname = cell.value
            if 'G' in cell.coordinate:
                date = cell.value
            if 'H' in cell.coordinate:
                days = cell.value
            if 'I' in cell.coordinate:
                wildcard = cell.value
                list_c.append(Challenge_Match(t_rank,g_rank,t_name,g_name,t_surname,g_surname,date,days,wildcard))
    workbook.active = 3
    worksheet = workbook.active
    if worksheet is None:
       worksheet = workbook.create_sheet()
    for row in worksheet.iter_rows(): #Διαβαζει τα στοιχεια των στατιστικων
        for cell in row:
            if 'A' in cell.coordinate:
                 t_rank = cell.value
            if 'B' in cell.coordinate:
                 t_name = cell.value
            if 'C' in cell.coordinate:
                t_surname = cell.value
            if 'D' in cell.coordinate:
                t_sets = cell.value
            if 'E' in cell.coordinate:
                g_rank = cell.value
            if 'F' in cell.coordinate:
                g_name = cell.value
            if 'G' in cell.coordinate:
                g_surname = cell.value
            if 'H' in cell.coordinate:
                g_sets = cell.value
            if 'I' in cell.coordinate:
                date = cell.value
                list_d.append(Statistics(t_rank, t_name, t_surname,t_sets, g_rank,g_name,g_surname,g_sets,date))
    return 
    
def write_excel(filename,list_a,list_b,list_c,list_d):#ΓΡΑΦΕΙ ΑΠΟ ΤΙΣ ΛΙΣΤΕΣ ΕΙΣΟΔΟΥ ΣΤΟ ΕΞΕΛ
    workbook = openpyxl.load_workbook(filename,read_only=False,keep_vba=False)
    sheets = workbook.sheetnames
    for i in range(len(sheets)):
        workbook.remove(workbook[sheets[i]])
    worksheet = workbook.create_sheet()
    workbook.active = 0
    worksheet = workbook.active
    
    for i in range(len(list_a)): # ΓΡΑΦΕΙ ΤΑ ΣΤΟΙΧΕΙΑ ΤΩΝ ΠΑΙΚΤΩΝ ΣΤΟ ΕΞΕΛ
        j = 1
        for value in list_a[i].__dict__.values():
            worksheet.cell(i+1,j).value = value
            j +=1
    worksheet = workbook.create_sheet()
    workbook.active = 1
    worksheet = workbook.active
    for i in range(len(list_b)): # ΓΡΑΦΕΙ ΤΑ ΣΤΟΙΧΕΙΑ ΤΩΝ ΜΗ ΕΓΚΥΡΩΝ ΜΑΤΣ ΠΡΟΚΛΗΣΕΩΝ ΣΤΟ ΕΞΕΛ
        j = 1
        for value in list_b[i].__dict__.values():
            worksheet.cell(i+1,j).value = value
            j+=1
    worksheet = workbook.create_sheet()
    workbook.active = 2
    worksheet = workbook.active
    
    for i in range(len(list_c)): # ΓΡΑΦΕΙ ΤΑ ΣΤΟΙΧΕΙΑ ΤΩΝ ΕΓΚΥΡΩΝ ΜΑΤΣ ΠΡΟΚΛΗΣΗΣ ΣΤΟ ΕΞΕΛ
        j=1
        for value in list_c[i].__dict__.values():
            worksheet.cell(i+1,j).value = value
            j+=1
    worksheet = workbook.create_sheet()
    workbook.active = 3
    worksheet = workbook.active
    for i in range(len(list_d)):
        j=1
        for value in list_d[i].__dict__.values():
            worksheet.cell(i+1,j).value = value
            j+=1
    workbook.save(filename)
    return
def write_config():#ΓΡΑΦΕΙ ΤΙΣ ΡΥΘΜΙΣΕΙΣ ΣΕ ΑΡΧΕΙΟ ΙΝΙ
    global WILDCARD
    global MAX_RANKING_CHALLENGE
    global POINT_SET_MATCH
    global filename_player_data
    global filename_out_ch
    global filename_valid_ch
    global filename_image
    
    config_object = ConfigParser()
    config_object["settings"] = {"WILDCARD":WILDCARD,"MAX_RANKING_CHALLENGE":MAX_RANKING_CHALLENGE,"POINT_SET_MATCH":POINT_SET_MATCH,"ranking": filename_ranking ,"background":filename_image}
    
    with  open("config.ini",'w') as conf:
        config_object.write(conf)
        conf.close
    return

def read_config():# ΔΙΑΒΑΖΕΙ ΤΙΣ ΡΥΘΜΙΣΕΙΣ ΑΠΟ ΑΡΧΕΙΟ ΙΝΙ
    global WILDCARD
    global MAX_RANKING_CHALLENGE
    global POINT_SET_MATCH
    global filename_ranking
    global filename_image
    if(os.path.exists(filename_conf)==True):
        config_object = ConfigParser()
        config_object.read("config.ini")
        settings = config_object["settings"]
        WILDCARD = int(settings["WILDCARD"])
        MAX_RANKING_CHALLENGE = int(settings["MAX_RANKING_CHALLENGE"])
        POINT_SET_MATCH = int(settings["POINT_SET_MATCH"])
        filename_ranking= settings['ranking']
        filename_image = settings['background']
    else:
        write_config()
        read_config()
    return

def startup():
    read_config()
    open_excel(filename_ranking,pl_data,ch_out,ch_valid,statistics)
    return
####ΚΛΑΣΗ ΓΙΑ ΤΟ LOG-IN####
class LOGIN(tkinter.Tk):
 def __init__(self):
    super().__init__()
    global filename_image
    self.closed = False
    self.title("Πρόγραμμα Κατάταξης Τέννις")
    self.resizable(False,False)
    self.bg = Image.open(filename_image)
    self.img = ImageTk.PhotoImage(self.bg)
    self.screen_width = self.winfo_screenwidth()
    self.screen_height = self.winfo_screenheight()
    self.w = self.img.width()
    self.h = self.img.height()
    self.geometry(f"1280x720+{self.screen_width//3}+{self.screen_height//3}")
    self.protocol("WM_DELETE_WINDOW",self.disable_event)
    canvas = tkinter.Canvas(self, width=1280, height=720,borderwidth=0,highlightthickness=0)
    canvas.pack(fill=BOTH,expand=YES)
    canvas.create_image(self.w,self.h,image = self.img)
    #canvas.create_text(1100,700,text="ΠΛΗ ΠΡΟ ΤΜΗΜΑ 50 ΟΜΑΔΑ 3",font=('Tahoma 18 bold'), fill='#922B21')
    canvas.create_text(640, 360, text="Χρήστης", font=('Tahoma 18 bold'), fill='#922B21')
    canvas.create_text(640, 460, text="Κωδικός", font=('Tahoma 18 bold'), fill='#922B21')
    self.user_entry = Entry(self, font=("Tahoma 18 bold"))
    self.user_entry.focus()
    canvas.create_window(640, 380, anchor="nw", window=self.user_entry)
    self.paswd = StringVar()
    self.password_entry = Entry(self, textvar=self.paswd, font=("Ariel 18 bold"), show="*")
    canvas.create_window(640, 480, anchor="nw", window=self.password_entry)
    login = Button(self, text="Έισοδος", font=("Tahoma 22 bold"),width=8, bg="grey", fg='#922B21', relief=RAISED, cursor="hand2", command=self.check,borderwidth=2).place(x=360, y=640)
    login = Button(self, text="Εξοδος", font=("Tahoma 22 bold"),width=8, bg="grey", fg='#922B21', relief=RAISED, cursor="hand2", command=self.destroy,borderwidth=2).place(x=720, y=640)
    canvas.create_window(640, 560, anchor="nw", window=login)

 def disable_event(self):
    pass

 def check(self):
    if  self.user_entry.get() == "":
        tkinter.messagebox.showinfo("Πρόγραμμα Κατάταξης Τέννις", "Παρακαλώ δώστε όνομα χρήστη")
    elif self.password_entry.get() == "":
        tkinter.messagebox.showinfo("Πρόγραμμα Κατάταξης Τέννις", "Παρακαλώ δώστε κώδικο")
    elif self.user_entry.get() == "" and self.password_entry.get() == "":
        tkinter.messagebox.showinfo("Πρόγραμμα Κατάταξης Τέννις", "Παρακαλώ δώστε όνομα χρήστη και κώδικο")
    elif self.user_entry.get() == "admin" and self.password_entry.get() == "1977":
        tkinter.messagebox.showinfo("Πρόγραμμα Κατάταξης Τέννις","Εισοδος Επιτυχής")
        self.paswd.set("")
        self.closed=True
        self.destroy()
    else:
        tkinter.messagebox.showinfo("Tennis Ladder", "Λάθος στοιχεία εισόδου")
        self.paswd.set("")
   

### MAIN###
 
if __name__ == '__main__':
    
    root = LOGIN()
    root.mainloop()
    if root.closed==True:
        root.destroy
        startup()
        root = MAIN()
        root.state("zoomed")
        root.print_ranking()
        root.mainloop()
