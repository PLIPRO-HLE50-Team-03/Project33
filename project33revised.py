### ΕΙΣΑΓΩΓΗ ###
import tkinter  
import tkinter.ttk
from tkinter import filedialog
import sys
from sys import *
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from configparser import ConfigParser
import os
import os.path

###########################ΜΕΤΑΒΛΗΤΕΣ ΠΡΟΓΡΑΜΜΑΤΟΣ####################################################
global WILDCARD  # ΔΗΛΩΝΕΙ ΠΟΣΑ ΜΑΤΣ ΠΡΕΠΕΙ ΝΑ ΠΑΙΞΕΙ Ο ΠΑΙΧΤΗΣ ΓΙΑ ΝΑ ΠΑΡΕΙ ΜΠΑΛΑΝΤΕΡ
WILDCARD = 3
global MAX_ACTIVE_CHALLENGES  # ΔΗΛΩΝΕΙ ΤΟΝ ΜΕΓΙΣΤΟ ΑΡΙΘΜΟ ΠΡΟΚΛΗΣΕΩΝ ΠΟΥ ΜΠΟΡΕΙ ΝΑ EXEI ENEPΓΕΣ  ΕΝΑΣ ΠΑΙΧΤΗΣ
MAX_ACTIVE_CHALLENGES = 2
global MAX_RANKING_CHALLENGE  # ΔΗΛΩΝΕΙ ΤΗΝ ΜΕΓΙΣΤΗ ΑΠΟΣΤΑΣΗ(ΔΙΑΦΟΡΑ ΚΑΤΑΤΑΞΗΣ) ΠΟΥ ΠΡΕΠΕΙ ΝΑ ΕΧΟΥΝΕ ΟΙ ΠΑΙΚΤΕΣ ΓΙΑ ΝΑ ΣΤΕΙΛΟΥΝΕ ΠΡΟΚΛΗΣΗ
MAX_RANKING_CHALLENGE = 3
global POINT_SET_MATCH # ΟΡΙΖΕΙ ΤΟΝ ΜΕΓΙΣΤΟ ΑΡΙΘΜΟ ΣΕΤ ΠΟΥ ΠΡΕΠΕΙ ΝΑ ΠΑΡΕΙ ΕΝΑΣ ΠΑΙΚΤΗΣ ΓΙΑ ΝΑ ΚΕΡΔΙΣΕΙ ΤΟ ΜΑΤΣ ΠΡΟΚΛΗΣΗΣ
POINT_SET_MATCH = 3
global filename_ranking # ΔΗΛΩΝΕΙ ΤΟ ΑΡΧΕΙΟ ΕXCEL ΠΟΥ ΣΩΖΟΝΤΑΙ ΤΑ ΣΤΟΙΧΕΙΑ ΤΩΝ ΠΑΙΚΤΩΝ
filename_ranking = os.path.abspath("ranking.xlsx") # ΤΟ ΑΡΧΕΙΟ ΜΕ ΤΗΝ ΚΑΤΑΤΑΞΗ
filename_conf = os.path.abspath("config.ini") # ΤΟ ΑΡΧΕΙΟ ΡΥΘΜΙΣΕΩΝ
#########################################################################################################
pl_data = [] # ΛΙΣΤΑ ΜΕ ΤΑ ΣΤΟΙΧΕΙΑ ΤΩΝ ΠΑΙΚΤΩΝ
ch_out = [] # ΛΙΣΤΑ ΜΕ ΤΙΣ ΜΗ ΕΠΙΒΕΒΑΙΩΜΕΝΕΣ ΠΡΟΚΛΗΣΕΙΣ
ch_valid = [] # ΛΙΣΤΑ ΜΕ ΤΙΣ ΕΓΚΥΡΕΣ ΠΡΟΚΛΗΣΕΙΣ
#########################################################################################################
##########################ΟΡΙΣΜΟΣ ΚΛΑΣΗΣ ΠΑΙΚΤΗΣ########################################################
class Player:
    def __init__(self,name,surname,age,m_pl,m_won,m_lost,sets_pl,sets_won,sets_lost,wildcard,active_ch_counter,active_ch):
        global WILDCARD
        global MAX_ACTIVE_CHALLENGES
        self.name = name    #ΟΝΟΜΑ
        self.surname = surname # ΕΠΩΝΥΜΟ
        self.age = age #ΗΛΙΚΙΑ
        self.m_pl = m_pl # ΜΑΤΣ ΣΥΝΟΛΟ
        self.m_won = m_won # ΚΕΡΔΙΣMΕΝΑ ΜΑΤΣ
        self.m_lost = m_lost # ΧΑΜΕΝΑ ΜΑΤΣ
        self.sets_pl = sets_pl # ΣΕΤ ΣΥΝΟΛΟ
        self.sets_won = sets_won # ΚΕΡΔΙΣMΕΝΑ ΣΕΤ
        self.sets_lost = sets_lost # ΧΑΜΕΝΑ ΣΕΤ
        self.wildcard = wildcard # ΤΟ ΑΝ ΕΧΕΙ ΜΠΑΛΑΝΤΕΡ
        self.active_ch_counter = active_ch_counter # ΑΡΙΘΜΟΣ ΠΡΟΚΛΗΣΕΩΝ ΔΕΝ ΜΠΟΡΕΙ ΑΝ ΥΠΕΡΒΕΙ ΤΟ MAX_ΑCTIVE_CHALLENGES
        self.active_ch = active_ch # ΤΟ ΕΑΝ ΜΠΟΡΕΙ ΝΑ ΔΕΚΤΕΙ Η ΝΑ ΔΩΣΕΙ ΠΡΟΚΛΗΣH
        return
    def increase_m_pl(self): # ΑΥΞΑΝΕΙ ΤΑ ΜΑΤΣ ΠΡΟΚΛΗΣΗΣ
        self.m_pl = self.m_pl+1
        return
    def increase_m_won(self): # ΑΥΞΑΝΕΙ ΤΑ ΜΑΤΣ ΠΟΥ ΠΑΙΧΤΗΚΑΝ ΣΥΝΟΛΙΚΑ
        self.m_won = self.m_won +1 
        return
    def increase_m_lost(self):#ΑΥΞΑΝΕΙ ΤΑ ΜΑΤΣ ΠΟΥ ΧΑΘΗΚΑΝ
        self.m_lost = self.m_lost +1
        return
    def increase_sets_pl(self):# ΑΥΞΑΝΕΙ ΤΑ ΣΕΤΣ ΠΟΥ ΠΑΙΧΤΗΚΑΝ ΣΥΝΟΛΙΚΑ
        self.sets_pl = self.sets_won +self.sets_lost
        return
    def increase_sets_won(self,number): # ΑΥΞΑΝΕΙ ΤΑ ΣΕΤΣ ΠΟΥ ΚΕΡΔΙΘΗΚΑΝ
        self.sets_won += number
        return
    def increase_sets_lost(self,number): # ΑΥΞΑΝΕΙ ΤΑ ΣΕΤΣ ΠΟΥ ΧΑΘΗΚΑΝ
        self.sets_lost += number
        return
    def change_wild_state(self): # ΑΛΛΑΖΕΙ ΤΗΝ ΚΑΤΑΣΤΑΣΗ ΤΟΥ WILDCARD TRUE <--> FALSE
        if(self.wildcard == False):
            self.wildcard = True
        if(self.wildcard == True):
            self.wildcard = False
        return
    def check_wild_state(self): # EΛΕΓΧΕΙ ΑΝ ΤΟ WILDCARD ΠΡΕΠΕΙ ΑΝ ΑΛΛΑΧΘΕΙ ΚΑΙ ΚΑΛΕΙ ΤΗΝ ΚΑΤΑΛΛΗΛΗ ΣΥΝΑΡΤΗΣΗ
        if(self.m_pl>=WILDCARD and self.m_pl % WILDCARD==0):
            self.change_wild_state()
        return  
    def increase_active_ch_counter(self): # ΑΥΞΑΝΕΙ ΤΟΝ ΑΡΙΘΜΟ ΤΩΝ ΠΡΟΚΛΗΣΕΩΝ ΠΟΥ ΕΧΕΙ ΔΩΣΕΙ Η ΔΕΚΤΕΙ Ο ΠΑΙΚΤΗΣ
        self.active_ch_counter += 1
        return
    def decrease_active_ch_counter(self): # ΜΕΙΩΝΕΙ ΤΟΝ ΑΡΙΘΜΟ ΤΩΝ ΠΡΟΚΛΗΣΕΩΝ ΠΟΥ ΕΧΕΙ ΔΩΣΕΙ Η ΔΕΚΤΕΙ Ο ΠΑΙΚΤΗΣ
        self.active_ch_counter -= 1
        return
    def check_active_challenges(self): # ΕΛΕΓΧΕΙ ΑΝ Ο ΑΡΙΘΜΟΣ ΤΩΝ ΠΡΟΚΛΗΣΕΩΝ ΠΟΥ ΕΧΟΥΝΕ ΔΟΘΕΙ ΞΕΠΕΡΝΑ ΤΟ ΜΑΧΙΜΟΥΜ ΚΑΙ ΑΛΛΑΖΕΙ ΤΗΝ ΚΑΤΑΣΤΑΣΗ 
        if(self.active_ch_counter<MAX_ACTIVE_CHALLENGES):
            self.active_ch = True
        else: self.active_ch =False
        return
        
    def __str__(self): # ΤΥΠΩΝΕΙ ΒΑΣΙΚΑ ΣΤΟΙΧΕΙΑ ΓΙΑ ΤΟΝ ΠΑΙΚΤΗ
        return f"{self.name}\t\t{self.surname}\t{self.age}\t{self.wildcard}\t{self.active_ch}"
            
    def __stats__(self): # ΤΥΠΩΝΕΙ ΑΝΑΛΑΥΤΙΚΑ ΣΤΟΙΧΕΙΑ ΓΙΑ ΤΟΝ ΠΑΙΚΤΗ
        return f"{self.name}\t{self.surname}\t\t{self.m_pl}\t{self.m_won}\t{self.m_lost}\t{self.sets_pl}\t{self.sets_won}\t{self.sets_lost}"
 ################ΟΡΙΣΜΟΣ ΚΛΑΣΗ ΜΑΤΣ ΠΡΟΚΛΗΣΗΣ###################
class Challenge_Match:
    def __init__(self,t_rank,g_rank,t_name,g_name,t_surname,g_surname):
        self.t_rank = t_rank    #ΚΑΤΑΤΑΞΗ ΠΑΙΚΤΗ ΣΤΟΧΟΥ
        self.g_rank = g_rank    #ΚΑΤΑΤΑΞΗ ΠΑΙΚΤΗ ΠΟΥ ΠΡΟΚΑΛΕΣΕ
        self.t_name = t_name    # ΟΝΟΜΑ ΣΤΟΧΟΥ
        self.g_name = g_name    # ΟΝΟΜΑ ΠΑΙΚΤΗ ΠΟΥ ΠΡΟΚΑΛΕΣΕ
        self.t_surname = t_surname  # ΕΠΙΘΕΤΟ ΠΑΙΚΤΗ ΣΤΟΧΟΥ
        self.g_surname = g_surname  # ΕΠΙΘΕΤΟ ΠΑΙΚΤΗ ΠΟΥ ΠΡΟΚΑΛΕΣΕ
        return
    def __str__(self): ## ΕΚΤΥΠΩΣΗ ΤΗΣ ΚΛΑΣΗΣ ΓΙΑ ΤΑ ΜΑΤΣ ΠΡΟΚΛΗΣΗΣ
        return f"{(self.t_rank+1)}\t{self.t_name}\t{self.t_surname}\tVS\t{self.g_rank+1}\t{self.g_name}\t{self.g_surname}" 

        
def make_player(list_a,name,surname,age): #ΔΗΜΙΟΥΡΓΕΙ ΤΟ ΑΝΤΙΚΕΙΜΕΝΟ ΤΗΣ ΚΛΑΣΗΣ ΠΑΙΚΤΗ ΚΑΙ ΤΟ ΑΠΟΘΗΚΕΥΕΙ ΣΕ ΛΙΣΤΑ
    list_a.append(Player(name,surname,age,m_pl=0,m_won=0,m_lost=0,sets_pl=0,sets_won=0,sets_lost=0,wildcard=True,active_ch_counter=0,active_ch=True))
    return  
def make_challenge(list_a,t_rank,g_rank,t_name,g_name,t_surname,g_surname):  #ΔΗΜΙΟΥΡΓΕΙ ΤΟ ΑΝΤΙΚΕΙΜΕΝΟ ΤΗΣ ΚΛΑΣΗΣ ΜΑΤΣ ΠΡΟΚΛΗΣΗΣ ΚΑΙ ΤΟ ΑΠΟΘΗΚΕΥΕΙ ΣΕ ΛΙΣΤΑ
    list_a.append(Challenge_Match(t_rank,g_rank,t_name,g_name,t_surname,g_surname))
    return    
def delete_object(list_a,index): #ΣΒΗΝΕΙ ΕΝΑ ΑΝΤΙΚΕΙΜΕΝΟ ΑΠΟ ΛΙΣΤΑ
    del list_a[index]
    return
def delete_ch_match(list_a,ranking): #ΣΒΗΝΕΙ ΕΝΑ ΑΝΤΙΚΕΙΜΕΝΟ ΑΠΟ ΤΗΝ ΛΙΣΤΑ 
    for i in range(len(list_a)):
        if(list_a[i].t_rank == ranking or list_a[i].g_rank == ranking):
            del list_a[i]
        return

def swap(list_a,winner,loser):# Η ΣΥΝΑΡΤΗΣΗ ΑΥΤΗ ΔΕΧΕΤΕ ΜΙΑ ΛΙΣΤΑ ΚΑΙ ΔΥΟ ΑΡΙΘΜΟΥΣ ΚΑΤΑΤΑΞΗΣ ΚΑΙ ΕΠΙΣΤΡΕΦΕΙ ΤΟΝ ΛΙΣΤΑ ΜΕ ΒΑΣΗ ΤΗΝ ΑΛΛΑΓΗ ΣΤΗΝ ΚΑΤΑΤΑΞΗ
    base = list_a[loser] # ΑΠΟΘΗΚΕΥΟΥΜΕ ΤΟΝ ΠΑΙΚΤΗ ΠΟΥ ΗΤΤΗΘΗΚΕ 
    list_a[loser] = list_a[winner] # ΣΤΗΝ ΘΕΣΗ ΤΟΥ ΗΤΤΗΜΕΝΟΥ ΠΕΡΝΑΕΙ Ο ΝΙΚΗΤΗΣ
    TEMP = []# ΤΟΠΙΚΕΣ ΜΕΤΑΒΛΗΤΕΣ ΠΟΥ ΣΩΖΟΥΝ ΤΑ ΣΤΟΙΧΕΙΑ ΤΗΣ ΛΙΣΤΑΣ
    temp = []
    for i in range(loser+1,len(list_a)):# ΕΠΑΝΑΛΗΨΗ ΑΠΟ ΤΗΝ ΘΕΣΗ ΤΟΥ ΗΤΤΗΜΕΝΟΥ ΑΥΞΗΜΕΝΗ ΚΑΤΑ ΜΙΑ ΘΕΣΗ ΕΩΣ ΤΟ ΤΕΛΟΣ ΤΙΣ ΛΙΣΤΑΣ
        if(i == loser+1): # ΕΑΝ ΤΟ Ι ΕΙΝΑΙ ΙΣΟ ΜΕ ΤΗΝ ΘΕΣΗ ΜΕΤΑ ΤΟΝ ΧΑΜΕΝΟ
           TEMP = list_a[i]
           temp = list_a[i]
           list_a[i] = base # ΠΕΡΝΑΜΕ ΤΟΝ ΧΑΜΕΝΟ ΜΙΑ ΘΕΣΗ ΠΙΟ ΚΑΤΩ
        if(i%2==0 and i !=loser+1): # ΤΟ ΙΔΙΟ ΚΑΝΟΥΜΕ ΚΑΙ ΓΙΑ ΚΑΘΕ ΑΛΛΗ ΘΕΣΗ
            TEMP = list_a[i]
            list_a[i] = temp
        if(i%2!=0 and i != loser+1):
            temp = list_a[i]
            list_a[i]=TEMP
    return
###################ΛΟΓΙΚΗ ΠΡΟΚΛΗΣΕΩΝ##############################################
###TRUE ΕΓΚΥΡΗ ΠΡΟΚΛΗΣΗ                                                          #
###FALSE ΜΗ ΕΓΚΥΡΗ                                                               #
### Η ΚΥΡΙΑ ΛΟΓΙΚΗ ΓΙΑ ΤΟ ΕΑΝ ΜΠΟΡΕΙ ΑΝ ΔΟΘΕΙ ΜΙΑ ΠΡΟΚΛΗΣΗ                       #
##################################################################################
'''def has_wildcard(t_name,t_surname,g_name,g_surname,list_a): #ΕΛΕΓΧΕΙ ΕΑΝ Ο ΠΑΙΚΤΗΣ ΕΧΕΙ WILDCARD(O ΠΑΙΚΤΗΣ ΠΟΥ ΘΕΛΕΙ ΝΑ ΔΩΣΕΙ ΠΡΟΚΛΗΣΗ) ΚΑΙ ΑΝ Ο ΠΑΙΚΤΗΣ ΑΠΟΔΕΚΤΗΣ ΜΠΟΡΕΙ ΝΑ ΔΕΚΤΕΙ ΑΛΛΗ ΠΡΟΚΛΗΣΗ
    target_name = t_name
    target_surname = t_surname
    giver_name = g_name
    giver_surname = g_surname
    list_player = list_a
    for i in range(0,len(list_player)):
        if(list_player[i].name == target_name and list_player[i].surname == target_surname):
            if(list_player[i].active_ch == True):
                for j in range(0,len(list_player)):
                    if(list_player[j].name == giver_name and list_player[j].surname == giver_surname):
                        if(list_player[j].wildcard == True and list_player[j].active_ch ==True):
                            print("HAS WILDCARD TRUE")
                            return True
        

    return False
def no_wildcard(t_name,t_surname,g_name,g_surname,list_b):# ΕΑΝ Ο ΠΑΙΚΤΗΣ ΔΕΝ ΕΧΕΙ ΜΠΑΛΑΝΤΕΡ ΑΛΛΑ ΜΠΟΡΕΙ ΝΑ ΔΩΣΕΙ ΠΡΟΚΛΗΣΗ ΚΑΙ Ο ΑΝΤΙΠΑΛΟΣ ΣΤΟΧΟΣ ΕΙΝΑΙ ΕΝΤΟΣ ΟΡΙΩΝ ΚΑΤΑΤΑΞΗΣ
    target_name = t_name
    target_surname = t_surname
    giver_name = g_name
    giver_surname = g_surname
    list_player = list_b
    for i in range(0,len(list_player)):
            if(list_player[i].name == target_name and list_player[i].surname == target_surname):
                if(list_player[i].active_ch == True):
                    for j in range(0,len(list_player)):
                        if(list_player[j].name == giver_name and list_player[j].surname == giver_surname):
                            if(i<j and abs(j - i)<=MAX_RANKING_CHALLENGE and list_player[j].active_ch == True and list_player[j].wildcard == False):
                                print("DOESNT HAVE WILDCARD TRUE")
                                return True
    return False
def duplicate(t_name,t_surname,g_name,g_surname,list_c): # ΕΛΕΓΧΕΙ ΑΝ Η ΠΡΟΚΛΗΣΗ ΕΧΕΙ ΗΔΗ ΔΟΘΕΙ (EINAI DUPLICATE)
    target_name = t_name
    target_surname = t_surname
    giver_name = g_name
    giver_surname = g_surname
    list_matches = list_c
    for i in range(0,len(list_matches)):
        if((list_matches[i].t_name == target_name and list_matches[i].t_surname == target_surname and list_matches[i].g_name == giver_name and list_matches[i].g_surname == giver_surname) ^ (list_matches[i].t_name == giver_name and list_matches[i].t_surname == giver_surname and list_matches[i].g_name == target_name and list_matches[i].g_surname == target_surname)):
            print("ITS A DUPLICATE CHALLENGE")
            return True        
    return False
def logic(t_name,t_surname,t_rank,g_name,g_surname,g_rank,list_a,list_b): # ΣΥΝΔΥΑΣΜΟΣ ΤΩΝ ΠΑΡΑΠΑΝΩ
    target_name = t_name
    target_surname = t_surname
    giver_name = g_name
    giver_surname = g_surname
    target_rank = t_rank
    giver_rank = g_rank
    list_player = list_a
    list_matches = list_b
    if((has_wildcard(target_name,target_surname,giver_name,giver_surname,list_player)==True and duplicate(target_name,target_surname,giver_name,giver_surname,list_matches)==False) ^ ((no_wildcard(target_name,target_surname,giver_name,giver_surname,list_player))==True and duplicate(target_name,target_surname,giver_name,giver_surname,list_matches)==False)):
         return True
    return False'''

def main_logic(t_name,t_surname,t_rank,g_name,g_surname,g_rank,list_a,list_b): # ΑΥΤΗ Η ΣΥΝΑΡΤΗΣΗ ΕΙΝΑΙ Η ΚΥΡΙΑ ΛΟΓΙΚΗ ΤΩΝ ΝΕΩΝ ΠΡΟΚΛΗΣΕΩΝ
    if(len(list_b)==0):
        if( g_rank>t_rank and abs(g_rank-t_rank)<=MAX_RANKING_CHALLENGE and list_a[g_rank].wildcard == False and list_a[g_rank].active_ch == True and list_a[t_rank].active_ch == True):
            return True
        if(g_rank>t_rank and list_a[g_rank].wildcard == True and list_a[g_rank].active_ch == True and list_a[t_rank].active_ch == True):
            return True
    else:
        for i in range(0,len(list_b)):
            if(list_b[i].t_name!= t_name and list_b[i].t_surname != t_surname and list_b[i].g_name != g_name and list_b[i].g_surname != g_surname or list_b[i].t_name!= g_name and list_b[i].t_surname != g_surname and list_b[i].g_name != t_name and list_b[i].g_surname != t_surname):
                if( g_rank>t_rank and abs(g_rank-t_rank)<=MAX_RANKING_CHALLENGE and list_a[g_rank].wildcard == False and list_a[g_rank].active_ch == True and list_a[t_rank].active_ch == True):
                    return True
                if(g_rank>t_rank  and list_a[g_rank].wildcard == True and list_a[g_rank].active_ch == True and list_a[t_rank].active_ch == True):
                    return True
    return False


####################################################################
### ΜΕΝΟΥ###########################################################
####################################################################
###ΑΥΤΗ Η ΚΛΑΣΗ ΕΙΝΑΙ ΤΟ ΚΥΡΙΟ ΜΕΝΟΥ###
class MAIN(tkinter.Tk):
    def __init__(self):
        super().__init__()
        self.title("Πρόγραμμα Κατάταξης Τέννις")
        self.screen_width = self.winfo_screenwidth()
        self.screen_height = self.winfo_screenheight()
        self.geometry(f"{self.screen_width}x{self.screen_height}")
        self.attributes('-fullscreen',True)
        self.configure(background='light cyan')
        self.text = tkinter.Text(height=40,width=140,font =("System", 10),bg='black',fg = 'white',relief=tkinter.RAISED)
        self.text.pack(anchor=tkinter.N,side=tkinter.TOP)
        self.terminate = tkinter.Button(self,height=1,width=6,text = "Εξοδος",activebackground="white",activeforeground="red",bd=8,relief=tkinter.RAISED,bg="red",fg="black",font=("System"),highlightcolor="black",command=self.terminate)
        self.terminate.pack(anchor = tkinter.S,side=tkinter.RIGHT)
        self.settings_button = tkinter.Button(self,height=1,width=9,text = "Ρυθμίσεις",activebackground="white",activeforeground="red",bd=8,relief=tkinter.RAISED,fg="black",bg="green",font=("System"),highlightcolor="black",command=self.settings)
        self.settings_button.pack(anchor = tkinter.S,side=tkinter.RIGHT)
        self.print_ranking_button = tkinter.Button(self,height=1,width=18,text ="Εκτύπωση Κατάταξης",activebackground="white",activeforeground="red",bd=8,relief=tkinter.RAISED,bg="silver",fg="black",font=("System"),highlightcolor="black",command=self.print_ranking)
        self.print_ranking_button.pack(anchor = tkinter.S,side=tkinter.LEFT)
        self.print_stats = tkinter.Button(self,height=1,width=19,text ="Εκτύπωση Στατιστικά",activebackground="white",activeforeground="red",bd=8,relief=tkinter.RAISED,bg="silver",fg="black",font=("System"),highlightcolor="black",command=self.print_ranking_stats)
        self.print_stats.pack(anchor = tkinter.S,side=tkinter.LEFT)
        self.print_out = tkinter.Button(self,height=1,width=17,text ="Εκτύπωση Εκρρεμών",activebackground="white",activeforeground="red",bd=8,relief=tkinter.RAISED,bg="silver",fg="black",font=("System"),highlightcolor="black",command=self.print_out_ch_matches)
        self.print_out.pack(anchor = tkinter.S,side=tkinter.LEFT)
        self.print_valid = tkinter.Button(self,height=1,width=16,text ="Εκτύπωση Ενεργών",activebackground="white",activeforeground="red",bd=8,relief=tkinter.RAISED,bg="silver",fg="black",font=("System"),highlightcolor="black",command=self.print_valid_ch_matches)
        self.print_valid.pack(anchor = tkinter.S,side=tkinter.LEFT)
        self.man_valid = tkinter.Button(self,height=1,width=18,text ="Ενημέρωση Ενεργών",activebackground="white",activeforeground="red",bd=8,relief=tkinter.RAISED,bg="yellow",fg="black",font=("System"),highlightcolor="black",command=self.ch_valid_man)
        self.man_valid.pack(anchor = tkinter.S,side=tkinter.LEFT)
        self.man_out = tkinter.Button(self,height=1,width=18,text ="Διαχείριση Εκρεμών",activebackground="white",activeforeground="red",bd=8,relief=tkinter.RAISED,bg="yellow",fg="black",font=("System"),highlightcolor="black",command=self.ch_out_man)
        self.man_out.pack(anchor = tkinter.S,side=tkinter.LEFT)
        self.ch_new_match = tkinter.Button(self,height=1,width=18,text ="Νέο Μάτς Πρόκλησης",activebackground="white",activeforeground="red",bd=8,relief=tkinter.RAISED,bg="yellow",fg="black",font=("System"),highlightcolor="black",command=self.new_ch_match)
        self.ch_new_match.pack(anchor = tkinter.S,side=tkinter.LEFT)
        self.del_pl = tkinter.Button(self,height=1,width=15,text ="Διαγραφή Παίχτη",activebackground="white",activeforeground="red",bd=8,relief=tkinter.RAISED,bg="blue",fg="black",font=("System"),highlightcolor="black",command=self.del_player)
        self.del_pl.pack(anchor = tkinter.S,side=tkinter.LEFT)
        self.new_pl = tkinter.Button(self,height=1,width=19,text ="Εγγραφή Νέου Παίχτη",activebackground="white",activeforeground="red",bd=8,relief=tkinter.RAISED,bg="blue",fg="black",font=("System"),highlightcolor="black",command=self.add_player)
        self.new_pl.pack(anchor = tkinter.S,side=tkinter.LEFT)
    
    ####ΚΟΥΜΠΙ ΤΕΡΜΑΤΙΣΜΟΥ####
    def terminate(self):
        global filename_player_data
        global filename_out_ch
        global filename_valid_ch
        write_excel(filename_ranking,pl_data,ch_out,ch_valid)
        root.destroy()
        
   
    ###ΤΥΠΩΝΕΙ ΤΗΝ ΒΑΣΙΚΗ ΚΑΤΑΤΑΞΗ###
    def print_ranking(self): 
        self.text.configure(state = 'normal')
        self.text.delete('1.0','end')
        self.text.insert(tkinter.INSERT,'Κατάταξη'+'\t\t'+'Ονομα'+'\t\t'+'Επώνυμο'+'\t'+'Ηλικία'+'\t'+'Wildcard'+'\t'+'Πρόκληση'+'\n')
        for index in range(len(pl_data)):
            self.text.insert(tkinter.INSERT,str(index+1)+'\t\t')
            self.text.insert(tkinter.INSERT,pl_data[index].__str__()+'\n')
        self.text.configure(state = 'disabled')
    
    ###TYΠΩΝΕΙ ΤΗΝ ΚΑΤΑΤΑΞΗ ΜΕ ΣΤΑΤΙΣΤΙΚΑ###
    def print_ranking_stats(self):
        self.text.configure(state = 'normal')
        self.text.delete('1.0','end')
        self.text.insert(tkinter.INSERT,'Κατάταξη'+'\t'+'Ονομα'+'\t'+'Επώνυμο'+'\t'+'   Παίχτηκαν'+'\t'+'    Νίκες'+'\t'+'      Ηττες'+'\t'+'   Σύνολο'+'\t'+'   Νικηφόρα'+'\t'+'   Χαμένα'+'\n')
        for index in range(len(pl_data)):
            self.text.insert(tkinter.INSERT,str(index+1)+'\t')
            self.text.insert(tkinter.INSERT,pl_data[index].__stats__()+'\n')
        self.text.configure(state = 'disabled')

    ###ΤΥΠΩΝΕΙ ΤΑ ΜΑΤΣ ΠΡΟΚΛΗΣΗΣ ΠΟΥ ΕΙΝΑΙ ΕΚΡΕΜΗ###
    def print_out_ch_matches(self):
        self.text.configure(state = 'normal')
        self.text.delete('1.0','end')
        for i in range(len(ch_out)):
             self.text.insert(tkinter.INSERT,"Πρόκληση"+'\t'+'\t'+str(i+1)+'η'+'\t' + ch_out[i].__str__()+'\n')
        self.left_text.configure(state = 'disabled')

    ###ΤΥΠΩΝΕΙ ΤΑ ΜΑΤΣ ΠΡΟΚΛΗΣΗΣ ΠΟΥ ΕΙΝΑΙ ΕΝΕΡΓΑ###
    def print_valid_ch_matches(self):
        self.text.configure(state='normal')
        self.text.delete('1.0','end')
        for i in range(len(ch_valid)):
            self.text.insert(tkinter.INSERT,"Πρόκληση"+'\t'+str(i+1)+'η'+'\t'+ch_valid[i].__str__()+'\n')
        self.text.configure(state = 'disabled')
    ###ΔΙΑΧΕΙΡΙΣΗ ΕΚΚΡΕΜΩΝ ΜΑΤΣ ΠΡΟΚΛΗΣΗΣ###
    def ch_out_man(self):
        top = OUTSTANDING_CHALLENGE_MATCH(self)
        top.grab_set()
    ###ΕΝΗΜΕΡΩΣΗ ΕΝΕΡΓΩΝ ΜΑΤΣ ΠΡΟΚΛΗΣΗΣ###
    def ch_valid_man(self):
        top = BRIEF_CHALLENGE_MATCH(self)
        top.grab_set()
    ###ΝΕΟ ΜΑΤΣ ΠΡΟΚΛΗΣΗΣ###
    def new_ch_match(self):
        top = NEW_CHALLENGE_MATCH(self)
        top.grab_set()
    ###ΔΙΑΓΡΑΦΗ ΠΑΙΚΤΗ###
    def del_player(self):
        top = DEL_PLAYER(self)
        top.grab_set()
    ###ΕΓΡΑΦΗ ΝΕΟΥ ΠΑΙΧΤΗ###
    def add_player(self):
        top = ADD_PLAYER(self)
        top.grab_set()
    ###ΡΥΘΜΙΣΕΙΣ###
    def settings(self):
        top = SETTINGS(self)
        top.grab_set()

###ΚΛΑΣΗ ΥΠΟ ΜΕΝΟΥ ΑΠΟΔΟΧΗ ΕΙΣΟΔΟΥ###

class CONFIRMED(tkinter.Toplevel): 
    def __init__(self,parent): 
        super().__init__(parent)
        self.title("Πρόγραμμα Κατάταξης Τέννις")
        self.screen_width = self.winfo_screenwidth()
        self.screen_height = self.winfo_screenheight()
        self.configure(background='light cyan')
        self.geometry(f"{self.screen_width//8}x{self.screen_height//8}+{self.screen_width//2}+{self.screen_height//2}")
        tkinter.Label(self,bg="silver",text="Αποδοχη",fg="green",font=("System",12)).pack()
        tkinter.Button(self,text='Επιστροφή',activebackground="white",activeforeground="red",fg="black",font=("System",10),highlightcolor="black",bg="green",padx=10,pady=10,command=self.destroy).pack(anchor=tkinter.S,side=tkinter.BOTTOM)

###ΚΛΑΣΗ ΥΠΟ ΜΕΝΟΥ ΑΠΟΡΡΙΨΗ ΕΙΣΟΔΟΥ###
class FAILED(tkinter.Toplevel): 
    def __init__(self,parent): 
        super().__init__(parent)
        self.title("Πρόγραμμα Κατάταξης Τέννις")
        self.screen_width = self.winfo_screenwidth()
        self.screen_height = self.winfo_screenheight()
        self.configure(background='light cyan')
        self.geometry(f"{self.screen_width//8}x{self.screen_height//8}+{self.screen_width//2}+{self.screen_height//2}")
        tkinter.Label(self,bg="red",text="Απόρριψη",fg="red",font=("System",12)).pack()
        tkinter.Button(self,text='Επιστροφή',activebackground="white",activeforeground="red",bg="blue",fg="black",font=("System",10),highlightcolor="black",padx=10,pady=10,command=self.destroy).pack(anchor=tkinter.S,side=tkinter.BOTTOM)

###ΚΛΑΣΗ ΥΠΟ ΜΕΝΟΥ ΕΓΓΡΑΦΗ ΝΕΟΥ ΠΑΙΚΤΗ###
class ADD_PLAYER(tkinter.Toplevel):
    def __init__(self,parent): 
        super().__init__(parent)
        self.title("Πρόγραμμα Κατάταξης Τέννις")
        self.screen_width = self.winfo_screenwidth()
        self.screen_height = self.winfo_screenheight()
        self.configure(background='light cyan')
        self.geometry(f"{self.screen_width//4}x{self.screen_height//4}+{self.screen_width//2}+{self.screen_height//2}")
        self.name = tkinter.StringVar()
        self.surname = tkinter.StringVar()
        self.age = tkinter.IntVar(value=16)
        tkinter.Label(self,text="Ονομα",bg="silver",fg="black",font=("System",12,"bold")).pack(anchor=tkinter.N,side=tkinter.TOP)
        tkinter.Entry(self,fg="black",bg="white",font=("System",12,"bold"),textvariable=self.name).pack(anchor=tkinter.N,side=tkinter.TOP)
        tkinter.Label(self,text="Επώνυμο",bg="silver",fg="black",font=("System",12,"bold")).pack(anchor=tkinter.N,side=tkinter.TOP)
        tkinter.Entry(self,fg="black",bg="white",font=("System",12,"bold"),textvariable=self.surname).pack(anchor=tkinter.N,side=tkinter.TOP)
        tkinter.Label(self,text="Ηλικία",bg="silver",fg="black",font=("System",12,"bold")).pack(anchor=tkinter.N,side=tkinter.TOP)
        tkinter.Entry(self,fg="black",bg="white",font=("System",12,"bold"),textvariable=self.age).pack(anchor=tkinter.N,side=tkinter.TOP)
        tkinter.Button(self,height=1,width=30,text = "Επιστροφή",activebackground="white",bg="blue",activeforeground="red",bd=8,relief=tkinter.RAISED,fg="black",font=("System",12),highlightcolor="black",command=self.destroy).pack(anchor=tkinter.SW,side=tkinter.LEFT)
        tkinter.Button(self,height=1,width=30,text = "Εγγραφή",activebackground="white",bg="green",activeforeground="red",bd=8,relief=tkinter.RAISED,fg="black",font=("System",12),highlightcolor="black",command=self.confirm).pack(anchor=tkinter.SW,side=tkinter.RIGHT)
###ΕΠΙΒΕΒΑΙΩΣΗ ΠΡΟΣΘΗΚΗΣ ΠΑΙΚΤΗ###
    def confirm(self):
        try:
                name=self.name.get()
                surname=self.surname.get()
                age=self.age.get()
        except:
                top = FAILED(self)
                top.grab_set()
                
        if(len(name)>=1 and len(surname)>=1 and age>=16):
            make_player(pl_data,name,surname,age) #ΠΡΟΣΘΕΤΕΙ ΤΟΝ ΠΑΙΧΤΗ ΕΑN 16+
            root.print_ranking()
            top = CONFIRMED(self)
            top.grab_set()
        
        

###ΚΛΑΣΗ ΥΠΟ ΜΕΝΟΥ ΔΙΑΓΡΑΦΗ ΠΑΙΚΤΗ####
class DEL_PLAYER(tkinter.Toplevel):
    def __init__(self,parent): 
        super().__init__(parent)
        self.title("Πρόγραμμα Κατάταξης Τέννις")
        self.screen_width = self.winfo_screenwidth()
        self.screen_height = self.winfo_screenheight()
        self.configure(background='light cyan')
        self.geometry(f"{self.screen_width//4}x{self.screen_height//4}+{self.screen_width//2}+{self.screen_height//2}")
        self.variable_del = tkinter.IntVar()
        tkinter.Label(self,text="Αριθμός Κατάταξης",bg="silver",fg="black",font=("System",12,"bold")).pack(anchor=tkinter.N,side=tkinter.TOP)
        tkinter.Entry(self,fg="black",bg="white",font=("System",12,"bold"),textvariable=self.variable_del).pack(anchor=tkinter.N,side=tkinter.TOP)
        tkinter.Button(self,height=1,width=30,bg="blue",text = "Επιστροφή",activebackground="white",activeforeground="red",bd=8,relief=tkinter.RAISED,fg="black",font=("System",12),highlightcolor="black",command=self.destroy).pack(anchor=tkinter.SW,side=tkinter.LEFT)
        tkinter.Button(self,height=1,width=30,bg="green",text = "Διαγραφή",activebackground="white",activeforeground="red",bd=8,relief=tkinter.RAISED,fg="black",font=("System",12),highlightcolor="black",command=self.confirm_del).pack(anchor=tkinter.SW,side=tkinter.RIGHT)
###EΠΙΒΕΒΑΙΩΣΗ ΔΙΑΓΡΑΦΗ ΠΑΙΚΤΗ###
    def confirm_del(self):
        try:
            var_del_player = self.variable_del.get()-1
        except:
            top = FAILED(self)
            top.grab_set()
               
        if(var_del_player >=0 and var_del_player<=(len(pl_data))):         
            delete_ch_match(ch_out,var_del_player)
            delete_ch_match(ch_valid,var_del_player)
            delete_object(pl_data,var_del_player)
            root.print_ranking()
            root.print_out_ch_matches()
            root.print_valid_ch_matches()
            top = CONFIRMED(self)
            top.grab_set()
        else:
            top = FAILED(self)
            top.grab_set()
               
###ΚΛΑΣΗ ΥΠΟΜΕΝΟΥ ΡΥΘΜΙΣΕΙΣ ##
class SETTINGS(tkinter.Toplevel):
    def __init__(self,parent):
        super().__init__(parent)
        self.title("Πρόγραμμα Κατάταξης Τέννις")
        self.screen_width = self.winfo_screenwidth()
        self.screen_height = self.winfo_screenheight()
        self.configure(background='light cyan')
        self.geometry(f"{self.screen_width//2}x{self.screen_height//2}+{self.screen_width//2}+{self.screen_height//2}")
        self.wildcard = tkinter.IntVar()
        self.max_ch_ranking = tkinter.IntVar()
        self.max_active = tkinter.IntVar() 
        self.max_sets = tkinter.IntVar() 
        tkinter.Label(self,bg="silver",text="Μάτς που παίζονται για Μπαλαντέρ",fg="black",font=("System",12,"bold")).pack(anchor=tkinter.N,side=tkinter.TOP)
        tkinter.Entry(self,bg="white",fg="black",font=("System",12,"bold"),textvariable=self.wildcard).pack(anchor=tkinter.N,side=tkinter.TOP)
        tkinter.Label(self,bg="silver",text="Απόσταση στην Κατάταξη μεταξύ παικτών για προκλήση",fg="black",font=("System",12,"bold")).pack(anchor=tkinter.N,side=tkinter.TOP)
        tkinter.Entry(self,bg="white",fg="black",font=("System",12,"bold"),textvariable=self.max_ch_ranking).pack(anchor=tkinter.N,side=tkinter.TOP)
        tkinter.Label(self,bg="silver",text="Mέγιστος αριθμός Επιτρεπώμενων Ενεργών Προκλήσεων",fg="black",font=("System",12,"bold")).pack(anchor=tkinter.N,side=tkinter.TOP)
        tkinter.Entry(self,bg="white",fg="black",font=("System",12,"bold"),textvariable=self.max_active).pack(anchor=tkinter.N,side=tkinter.TOP)
        tkinter.Label(self,bg="silver",text="Aριθμός σετ που πρέπει να κερδίσει ο παίκτης για να πάρει το παιχνίδι",fg="black",font=("System",12,"bold")).pack(anchor=tkinter.N,side=tkinter.TOP)
        tkinter.Entry(self,bg="white",fg="black",font=("System",12,"bold"),textvariable=self.max_sets).pack(anchor=tkinter.N,side=tkinter.TOP)
        tkinter.Button(self,height=1,width=10,bg="blue",text = "Επιστροφή",activebackground="white",activeforeground="red",bd=8,relief=tkinter.RAISED,fg="black",font=("System",12),highlightcolor="black",command=self.destroy).pack(anchor=tkinter.SW,side=tkinter.LEFT)
        tkinter.Button(self,height=1,width=10,bg="green",text = "Εγγραφή",activebackground="white",activeforeground="red",bd=8,relief=tkinter.RAISED,fg="black",font=("System",12),highlightcolor="black",command=self.define_settings).pack(anchor=tkinter.SW,side=tkinter.RIGHT)
### ΡΥΘΜΙΣΕΙΣ###
    def define_settings(self):
        global WILDCARD
        global MAX_RANKING_CHALLENGE
        global MAX_ACTIVE_CHALLENGES
        global POINT_SET_MATCH
        try:
            WILDCARD = self.wildcard.get()
            MAX_RANKING_CHALLENGE = self.max_ch_ranking.get()
            MAX_ACTIVE_CHALLENGES = self.max_active.get()
            POINT_SET_MATCH = self.max_sets.get()
            write_config()
            top = CONFIRMED(self)
            top.grab_set()
        except:
            top = FAILED(self)
            top.grab_set()
        return 
    
    '''def load_from_file(self):
        global filename_ranking
        filename_ranking = tkinter.filedialog.askopenfilename(title='Αρχείο Κατάταξης')
        if(filename_ranking == ""):
            filename_ranking = os.path.abspath("ranking.xlsx")
        open_excel(filename_ranking,pl_data,ch_out,ch_valid)
        return'''
       
    
###ΚΛΑΣΗ ΥΠΟΜΕΝΟΥ ΝΕΟ ΜΑΤΣ ΠΡΟΚΛΗΣΗΣ###
class NEW_CHALLENGE_MATCH(tkinter.Toplevel):
    def __init__(self,parent):
        super().__init__(parent)
        self.title("Πρόγραμμα Κατάταξης Τέννις")
        self.screen_width = self.winfo_screenwidth()
        self.screen_height = self.winfo_screenheight()
        self.configure(background='light cyan')
        self.geometry(f"{self.screen_width//4}x{self.screen_height//4}+{self.screen_width//2}+{self.screen_height//2}")
        self.challenger = tkinter.IntVar()
        self.champion = tkinter.IntVar()
        tkinter.Label(self,bg="silver",text="Κατάταξη Παίκτη που προκαλεί",fg="black",font=("System",12,"bold")).pack(anchor=tkinter.N,side=tkinter.TOP)
        tkinter.Entry(self,bg="white",fg="black",font=("System",12,"bold"),textvariable=self.challenger).pack(anchor=tkinter.N,side=tkinter.TOP)
        tkinter.Label(self,bg="silver",text="Κατάταξη Παίκτη αποδέκτης πρόκλησης",fg="black",font=("System",12,"bold")).pack(anchor=tkinter.N,side=tkinter.TOP)
        tkinter.Entry(self,bg="white",fg="black",font=("System",12,"bold"),textvariable=self.champion).pack(anchor=tkinter.N,side=tkinter.TOP)
        tkinter.Button(self,bg="blue",height=1,width=10,text = "Επιστροφή",activebackground="white",activeforeground="red",bd=8,relief=tkinter.RAISED,fg="black",font=("System",12),highlightcolor="black",command=self.destroy).pack(anchor=tkinter.SW,side=tkinter.LEFT)
        tkinter.Button(self,bg="green",height=1,width=10,text = "Εγγραφή",activebackground="white",activeforeground="red",bd=8,relief=tkinter.RAISED,fg="black",font=("System",12),highlightcolor="black",command=self.confirm_ch_match).pack(anchor=tkinter.SW,side=tkinter.RIGHT)
    def confirm_ch_match(self):
        try:
            g_rank = self.challenger.get()-1
            t_rank = self.champion.get()-1          
        except:
            top = FAILED(self)
            top.grab_set()
        if( g_rank >=0 and g_rank <= (len(pl_data))  and t_rank >=0 and t_rank <=len(pl_data) and g_rank>t_rank):
           ### ΕΛΕΓΧΕΙ ΑΝ ΟΙ ΑΡΙΘΜΟΙ ΕΙΝΑΙ ΣΤΗΝ ΕΠΙΤΡΕΠΟΜΕΝΗ ΑΠΟΣΤΑΣΗ
            if(main_logic(pl_data[t_rank].name,pl_data[t_rank].surname,t_rank,pl_data[g_rank].name,pl_data[g_rank].surname,g_rank,pl_data,ch_out)==True and pl_data[g_rank].active_ch==True and pl_data[t_rank].active_ch == True and pl_data[g_rank].active_ch_counter<MAX_ACTIVE_CHALLENGES and pl_data[t_rank].active_ch_counter<MAX_ACTIVE_CHALLENGES):
                pl_data[g_rank].increase_active_ch_counter()
                pl_data[g_rank].check_active_challenges()
                pl_data[t_rank].increase_active_ch_counter()
                pl_data[t_rank].check_active_challenges()
                if(abs(g_rank-t_rank)>=MAX_RANKING_CHALLENGE and pl_data[g_rank].wildcard==True):
                    pl_data[g_rank].change_wild_state()
                make_challenge(ch_out,t_rank,g_rank,pl_data[t_rank].name,pl_data[g_rank].name,pl_data[t_rank].surname,pl_data[g_rank].surname)
                root.print_out_ch_matches()
                top = CONFIRMED(self)
                top.grab_set()
                    
                    
            else:
                top = FAILED(self)
                top.grab_set()
                  
        else:
            top = FAILED(self)
            top.grab_set()          

###ΚΛΑΣΗ ΥΠΟΜΕΝΟΥ ΕΝΗΜΕΡΩΣΗ ΕΝΕΡΓΩΝ ΜΑΤΣ ΠΡΟΚΛΗΣΗΣ###
class BRIEF_CHALLENGE_MATCH(tkinter.Toplevel):
    def __init__(self,parent):
        super().__init__(parent)
        self.title("Πρόγραμμα Κατάταξης Τέννις")
        self.screen_width = self.winfo_screenwidth()
        self.screen_height = self.winfo_screenheight()
        self.configure(background='light cyan')
        self.geometry(f"{self.screen_width//2}x{self.screen_height//2}+{self.screen_width//2}+{self.screen_height//2}")
        self.serial = tkinter.IntVar()
        self.winner = tkinter.IntVar()
        self.loser = tkinter.IntVar()
        self.sets_winner = tkinter.IntVar()
        self.sets_loser = tkinter.IntVar()
        tkinter.Label(self,bg="silver",text="Αριθμός μάτς κατάταξης προς ενημέρωση",fg="black",font=("System",12,"bold")).pack(anchor=tkinter.N,side=tkinter.TOP)
        tkinter.Entry(self,bg="white",fg="black",font=("System",12,"bold"),textvariable=self.serial).pack(anchor=tkinter.N,side=tkinter.TOP)
        tkinter.Label(self,bg="silver",text="Αριθμός κατάταξης παίχτη που νίκησε",fg="black",font=("System",12,"bold")).pack(anchor=tkinter.N,side=tkinter.TOP)
        tkinter.Entry(self,bg="white",fg="black",font=("System",12,"bold"),textvariable=self.winner).pack(anchor=tkinter.N,side=tkinter.TOP)
        tkinter.Label(self,bg="silver",text="Σετ που πήρε ο Νικητής",fg="black",font=("System",12,"bold")).pack(anchor=tkinter.N,side=tkinter.TOP)
        tkinter.Entry(self,bg="white",fg="black",font=("System",12,"bold"),textvariable=self.sets_winner).pack(anchor=tkinter.N,side=tkinter.TOP)
        tkinter.Label(self,bg="silver",text="Αριθμός κατάταξης παίχτη που ηττήθηκε",fg="black",font=("System",12,"bold")).pack(anchor=tkinter.N,side=tkinter.TOP)
        tkinter.Entry(self,bg="white",fg="black",font=("System",12,"bold"),textvariable=self.loser).pack(anchor=tkinter.N,side=tkinter.TOP)
        tkinter.Label(self,bg="silver",text="Σετ που πήρε ο Ηττημένος",fg="black",font=("System",12,"bold")).pack(anchor=tkinter.N,side=tkinter.TOP)
        tkinter.Entry(self,bg="white",fg="black",font=("System",12,"bold"),textvariable=self.sets_loser).pack(anchor=tkinter.N,side=tkinter.TOP)
        tkinter.Button(self,bg="blue",height=1,width=10,text = "Επιστροφή",activebackground="white",activeforeground="red",bd=8,relief=tkinter.RAISED,fg="black",font=("System",12),highlightcolor="black",command=self.destroy).pack(anchor=tkinter.SW,side=tkinter.LEFT)
        tkinter.Button(self,bg="green",height=1,width=10,text = "Εγγραφή",activebackground="white",activeforeground="red",bd=8,relief=tkinter.RAISED,fg="black",font=("System",12),highlightcolor="black",command=self.brief_match).pack(anchor=tkinter.SW,side=tkinter.RIGHT)
    def brief_match(self):
        
        serial = self.serial.get()-1
        winner = self.winner.get()-1
        loser = self.loser.get()-1
        sets_winner = self.sets_winner.get()
        sets_loser = self.sets_loser.get()        
        if( sets_winner>=1 and sets_winner<=POINT_SET_MATCH and sets_loser>=1 and sets_loser<=POINT_SET_MATCH and sets_winner>sets_loser and winner>=0 and winner<=len(pl_data) and loser>=0 and loser<=len(pl_data) and serial>=0 and serial<=(len(ch_valid))):#EΑΝ ΟΙ ΑΡΙΘΜΟΙ ΥΠΟΚΕΙΝΤΑΙ ΣΤΟΥΣ ΚΑΝΟΝΕΣ ΓΙΑ ΤΑ ΠΑΙΧΝΔΙΑ
            pl_data[winner].increase_m_pl()
            pl_data[loser].increase_m_pl()
            pl_data[winner].increase_m_won()
            pl_data[loser].increase_m_lost()
            pl_data[winner].increase_sets_won(sets_winner)
            pl_data[winner].increase_sets_lost(sets_loser)
            pl_data[loser].increase_sets_won(sets_loser)
            pl_data[loser].increase_sets_lost(sets_winner)
            pl_data[winner].increase_sets_pl()
            pl_data[loser].increase_sets_pl()
            pl_data[winner].decrease_active_ch_counter()
            pl_data[loser].decrease_active_ch_counter()
            pl_data[winner].check_active_challenges()
            pl_data[loser].check_active_challenges()
            swap(pl_data,winner,loser)
            delete_object(ch_valid,serial)
            root.print_valid_ch_matches()
            root.print_ranking()
            top = CONFIRMED(self)
            top.grab_set() 
        else:
            top = FAILED(self)
            top.grab_set()
                
###ΚΛΑΣΗ ΥΠΟΜΕΝΟΥ ΔΙΑΧΕΙΡΙΣΗ ΕΚΡΕΜΜΩΝ ΜΑΤΣ ΠΡΟΚΛΗΣΗΣ###
class OUTSTANDING_CHALLENGE_MATCH(tkinter.Toplevel):
    def __init__(self,parent):
        super().__init__(parent)
        self.title("Πρόγραμμα Κατάταξης Τέννις")
        self.screen_width = self.winfo_screenwidth()
        self.screen_height = self.winfo_screenheight()
        self.configure(background='light cyan')
        self.geometry(f"{self.screen_width//4}x{self.screen_height//4}+{self.screen_width//2}+{self.screen_height//2}")
        self.match = tkinter.IntVar()
        tkinter.Label(self,bg="silver",text="Αριθμός εκρεμούς μάτς κατάταξης",fg="black",font=("System",12,"bold")).pack(anchor=tkinter.N,side=tkinter.TOP)
        tkinter.Entry(self,bg="white",fg="black",font=("System",12,"bold"),textvariable=self.match).pack(anchor=tkinter.N,side=tkinter.TOP)
        tkinter.Button(self,bg="blue",height=1,width=10,text = "Επιστροφή",activebackground="white",activeforeground="red",bd=8,relief=tkinter.RAISED,fg="black",font=("System",12),highlightcolor="black",command=self.destroy).pack(anchor=tkinter.SW,side=tkinter.LEFT)
        tkinter.Button(self,bg="green",height=1,width=10,text = "Αποδοχή",activebackground="white",activeforeground="red",bd=8,relief=tkinter.RAISED,fg="black",font=("System",12),highlightcolor="black",command=self.confirm_match).pack(anchor=tkinter.SW,side=tkinter.RIGHT)
        tkinter.Button(self,bg="red",height=1,width=10,text = "Απόρριψη",activebackground="white",activeforeground="red",bd=8,relief=tkinter.RAISED,fg="black",font=("System",12),highlightcolor="black",command=self.deny_match).pack(anchor=tkinter.SW,side=tkinter.RIGHT)
    ###ΑΠΟΔΟΧΗ ΕΚΡΕΜΜΟΥΣ ΜΑΤΣ ΠΡΟΚΛΗΣΗΣ###
    def confirm_match(self):
        try:
            number = self.match.get() - 1
        except:
            pop_up = FAILED(self)
            pop_up.grab_set()
        if(number>=0 and number <=(len(ch_out))):
            t_rank = ch_out[number].t_rank
            g_rank = ch_out[number].g_rank
            if(pl_data[t_rank].active_ch==True):
                make_challenge(ch_valid,t_rank,g_rank,pl_data[t_rank].name,pl_data[g_rank].name,pl_data[t_rank].surname,pl_data[g_rank].surname)             
                delete_object(ch_out,number)
                root.print_out_ch_matches()
                root.print_valid_ch_matches()
                pop_up = CONFIRMED(self)
                pop_up.grab_set()
            else:
                    pop_up = FAILED(self)
                    pop_up.grab_set()
        else:
                pop_up = FAILED(self)
                pop_up.grab_set()

    ###ΑΡΝΗΣΗ ΕΚΡΕΜΟΥΣ ΜΑΤΣ ΠΡΟΚΛΗΣΗΣ###
    def deny_match(self):
        try:
            number = self.match.get() -1
        except:
            pop_up = FAILED(self)
            pop_up.grab_set()
        if(number>=0 and number<=len(ch_out)):
            pl_data[ch_out[number].g_rank].decrease_active_ch_counter()
            pl_data[ch_out[number].g_rank].check_active_challenges()
            delete_object(ch_out,number)
            root.print_out_ch_matches()
            pop_up = CONFIRMED(self)
            pop_up.grab_set()
        else:
            pop_up = FAILED(self)
            pop_up.grab_set()
               
               
def open_excel(filename,list_a,list_b,list_c):
    try:
        workbook = openpyxl.load_workbook(filename)
        workbook.active = 0
        worksheet = workbook.active
    except:
        workbook = Workbook()
        worksheet = workbook.active
        workbook.save('ranking.xlsx')
    if worksheet is None:
        worksheet = workbook.create_sheet()
    for row in worksheet.iter_rows(): # ΔΙΑΒΑΖΕΙ ΤΑ ΣΤΟΙΧΕΙΑ ΤΩΝ ΠΑΙΚΤΩΝ
        for cell in row:
            if 'A' in cell.coordinate:
                name = cell.value
            if 'B' in cell.coordinate:
                surname = cell.value
            if 'C' in cell.coordinate:
                age = cell.value
            if 'D' in cell.coordinate:    
                m_pl = cell.value
            if 'E' in cell.coordinate:        
                m_won = cell.value
            if 'F' in cell.coordinate:
                m_lost = cell.value
            if 'G' in cell.coordinate:
                sets_pl = cell.value
            if 'H' in cell.coordinate:
                sets_won = cell.value
            if 'I' in cell.coordinate:
                sets_lost = cell.value
            if 'J' in cell.coordinate:
                wildcard = cell.value
            if 'K' in cell.coordinate:
                active_ch_counter = cell.value
            if 'L' in cell.coordinate:
                active_ch = cell.value
        list_a.append(Player(name,surname,age,m_pl,m_won,m_lost,sets_pl,sets_won,sets_lost,wildcard,active_ch_counter,active_ch))
    workbook.active = 1
    worksheet = workbook.active
    if worksheet is None:
        worksheet = workbook.create_sheet()
        workbook.save('ranking.xlsx')

    for row in worksheet.iter_rows(): # ΔΙΑΒΑΖΕΙ ΤΑ ΣΤΟΙΧΕΙΑ ΤΩΝ ΜΗ  ΕΓΚΥΡΩΝ ΜΑΤΣ ΠΡΟΚΛΗΣΗΣ
        for cell in row:
            if 'A' in cell.coordinate:
                t_rank = cell.value
            if 'B' in cell.coordinate:
                g_rank = cell.value
            if 'C' in cell.coordinate:
                t_name = cell.value
            if 'D' in cell.coordinate:
                g_name = cell.value
            if 'E' in cell.coordinate:
                t_surname = cell.value
            if 'F' in cell.coordinate:
                g_surname = cell.value
        list_b.append(Challenge_Match(t_rank,g_rank,t_name,g_name,t_surname,g_surname))
    workbook.active = 2
    worksheet = workbook.active
    if worksheet is None:
        worksheet = workbook.create_sheet()
        workbook.save('ranking.xlsx')
    
    for row in worksheet.iter_rows(): # ΔΙΑΒΑΖΕΙ ΤΑ ΣΤΟΙΧΕΙΑ ΤΩΝ ΕΓΚΥΡΩΝ ΜΑΤΣ ΠΡΟΚΛΗΣΗΣ
        for cell in row:
            if 'A' in cell.coordinate:
                t_rank = cell.value
            if 'B' in cell.coordinate:
                g_rank = cell.value
            if 'C' in cell.coordinate:
                t_name = cell.value
            if 'D' in cell.coordinate:
                g_name = cell.value
            if 'E' in cell.coordinate:
                t_surname = cell.value
            if 'F' in cell.coordinate:
                g_surname = cell.value
        list_c.append(Challenge_Match(t_rank,g_rank,t_name,g_name,t_surname,g_surname))

    return 
    
def write_excel(filename,list_a,list_b,list_c):
    workbook = openpyxl.load_workbook(filename,read_only=False,keep_vba=False)
    sheets = workbook.sheetnames
    for i in range(len(sheets)):
        workbook.remove(workbook[sheets[i]])
    worksheet = workbook.create_sheet()
    workbook.active = 0
    worksheet = workbook.active
    temp = []
    
    for i in range(len(list_a)): # ΓΡΑΦΕΙ ΤΑ ΣΤΟΙΧΕΙΑ ΤΩΝ ΠΑΙΚΤΩΝ ΣΤΟ ΕΞΕΛ
        j = 1
        for value in list_a[i].__dict__.values():
            worksheet.cell(i+1,j).value = value
            j+=1
    worksheet = workbook.create_sheet()
    workbook.active = 1
    worksheet = workbook.active
    
    for i in range(len(list_b)): # ΓΡΑΦΕΙ ΤΑ ΣΤΟΙΧΕΙΑ ΤΩΝ ΜΗ ΕΓΚΥΡΩΝ ΜΑΤΣ ΠΡΟΚΛΗΣΕΩΝ ΣΤΟ ΕΞΕΛ
        j = 1
        for value in list_b[i].__dict__.values():
            worksheet.cell(i+1,j).value = value
            j+=1
    worksheet = workbook.create_sheet()
    workbook.active = 2
    worksheet = workbook.active
    
    for i in range(len(list_c)): # ΓΡΑΦΕΙ ΤΑ ΣΤΟΙΧΕΙΑ ΤΩΝ ΕΓΚΥΡΩΝ ΜΑΤΣ ΠΡΟΚΛΗΣΗΣ ΣΤΟ ΕΞΕΛ
        j=1
        for value in list_c[i].__dict__.values():
            worksheet.cell(i+1,j).value = value
            j+=1
    workbook.save(filename)
    return
def write_config():#ΓΡΑΦΕΙ ΤΙΣ ΡΥΘΜΙΣΕΙΣ ΣΕ ΑΡΧΕΙΟ ΙΝΙ
    global WILDCARD
    global MAX_ACTIVE_CHALLENGES
    global MAX_RANKING_CHALLENGE
    global POINT_SET_MATCH
    global filename_player_data
    global filename_out_ch
    global filename_valid_ch
    
    config_object = ConfigParser()
    config_object["settings"] = {"WILDCARD":WILDCARD,"MAX_ACTIVE_CHALLENGES":MAX_ACTIVE_CHALLENGES,"MAX_RANKING_CHALLENGE":MAX_RANKING_CHALLENGE,"POINT_SET_MATCH":POINT_SET_MATCH,"ranking": filename_ranking }
    
    with  open("config.ini",'w') as conf:
        config_object.write(conf)
        conf.close
    return

def read_config():# ΔΙΑΒΑΖΕΙ ΤΙΣ ΡΥΘΜΙΣΕΙΣ ΑΠΟ ΑΡΧΕΙΟ ΙΝΙ
    global WILDCARD
    global MAX_ACTIVE_CHALLENGES
    global MAX_RANKING_CHALLENGE
    global POINT_SET_MATCH
    global filename_ranking
    if(os.path.exists(filename_conf)==True):
        config_object = ConfigParser()
        config_object.read("config.ini")
        settings = config_object["settings"]
        WILDCARD = int(settings["WILDCARD"])
        MAX_ACTIVE_CHALLENGES = int(settings["MAX_ACTIVE_CHALLENGES"])
        MAX_RANKING_CHALLENGE = int(settings["MAX_RANKING_CHALLENGE"])
        POINT_SET_MATCH = int(settings["POINT_SET_MATCH"])
        filename_ranking= settings['ranking']
    else:
        write_config()
        read_config()
    return

def startup():
    read_config()
    open_excel(filename_ranking,pl_data,ch_out,ch_valid)
    return
### MAIN###
 
if __name__ == '__main__':
    startup()
    root = MAIN()
    root.print_ranking()
    root.mainloop()